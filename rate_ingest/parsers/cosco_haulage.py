from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from rate_ingest.models import (
    ParserTemplate,
    RateCard,
    RateChargeLine,
    RateImport,
    RateNote,
    RateOffer,
)
from rate_ingest.normalize import normalize_text


HEADER_LABELS = {"load location", "pol", "bound", "transport mode", "40ft/hc usd"}


def parse_workbook(
    path: Path,
    template: ParserTemplate,
    rate_import: RateImport,
) -> tuple[RateCard, list[RateOffer], list[RateChargeLine], list[RateNote]]:
    card = RateCard(
        rate_import_id=rate_import.id,
        provider_name=template.provider_name,
        carrier_name=template.defaults.get("carrier_name", template.provider_name),
        document_type=template.document_type,
        commodity=template.defaults.get("commodity"),
        currency_default=template.defaults.get("currency_default", "USD"),
        all_in_flag=False,
        notes_summary="COSCO origin haulage tariffs by load location and POL.",
    )
    offers: list[RateOffer] = []
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header_index = None
            header: list[str] = []
            for index, row in enumerate(rows):
                candidate = [normalize_text(value).casefold() for value in row]
                if HEADER_LABELS.issubset(set(candidate)):
                    header_index = index
                    header = candidate
                    break
            if header_index is None:
                continue
            indexes = {label: header.index(label) for label in HEADER_LABELS}
            blank_rows = 0
            for row_index, row in enumerate(rows, start=header_index + 2):
                values = list(row)
                if not any(normalize_text(value) for value in values):
                    blank_rows += 1
                    if offers and blank_rows >= 50:
                        break
                    continue
                blank_rows = 0
                location = value_at(values, indexes["load location"])
                port_raw = value_at(values, indexes["pol"])
                bound = normalize_text(value_at(values, indexes["bound"])).upper()
                amount = to_float(value_at(values, indexes["40ft/hc usd"]))
                if not location or not port_raw or amount is None or amount <= 0:
                    continue
                if bound and bound not in {"OB", "OUTBOUND", "EXPORT"}:
                    continue
                port = canonical_port(port_raw)
                offers.append(
                    RateOffer(
                        rate_card_id=card.id,
                        offer_reference=f"{sheet.title}:{port}",
                        commodity=card.commodity,
                        origin=location,
                        place_of_receipt=location,
                        pol=port,
                        pod=port,
                        final_destination=port,
                        equipment_type=template.defaults.get("equipment_type", "40HC"),
                        service_mode="Door -> CY",
                        base_amount=round(amount, 2),
                        base_currency=card.currency_default,
                        all_in_flag=False,
                        raw_sheet_name=sheet.title,
                        raw_row_reference=f"{sheet.title}!R{row_index}",
                        raw_row_json={
                            "collection_place": location,
                            "pol_raw": normalize_text(port_raw),
                            "bound": bound,
                            "transport_mode": normalize_text(
                                value_at(values, indexes["transport mode"])
                            ),
                            "source_file_name": path.name,
                        },
                    )
                )
    finally:
        workbook.close()

    notes = [
        RateNote(
            rate_card_id=card.id,
            note_type="commercial",
            note_text="COSCO haulage is kept separate and added to quay-to-quay quotes at quote time.",
            source_reference=path.name,
        )
    ]
    return card, offers, [], notes


def value_at(row: list[object], index: int) -> object | None:
    return row[index] if index < len(row) else None


def to_float(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        text = normalize_text(value).replace(",", "")
        try:
            return float(text)
        except ValueError:
            return None


def canonical_port(raw: object) -> str:
    text = normalize_text(raw).upper()
    if "FELIXSTOWE" in text or text == "GBFXT":
        return "Felixstowe"
    if "SOUTHAMPTON" in text or text == "GBSOU":
        return "Southampton"
    return normalize_text(raw)
