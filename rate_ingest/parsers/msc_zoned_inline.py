from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

from rate_ingest.models import ParserTemplate, RateCard, RateChargeLine, RateImport, RateNote, RateOffer
from rate_ingest.normalize import normalize_equipment, normalize_text, parse_amount, parse_date_range, parse_date_value


def parse_workbook(
    path: Path,
    template: ParserTemplate,
    rate_import: RateImport,
) -> tuple[RateCard, list[RateOffer], list[RateChargeLine], list[RateNote]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    rules = template.inline_haulage_rules
    metadata = extract_cover_metadata(workbook)
    card_valid_from, card_valid_to = parse_compact_date_range(metadata.get("Validity From/To"))
    card = RateCard(
        rate_import_id=rate_import.id,
        provider_name=template.provider_name,
        carrier_name=template.provider_name,
        document_type=template.document_type,
        commodity=metadata.get("Commodity") or template.defaults.get("commodity"),
        currency_default=template.defaults.get("currency_default", "GBP"),
        valid_from=card_valid_from,
        valid_to=card_valid_to,
        all_in_flag=False,
        notes_summary="MSC door-to-quay rate selected by city, POL, and zone; documentation is additional.",
    )

    haulage_rows = load_haulage_rows(workbook, rules)
    ambiguous_cities = find_ambiguous_cities(haulage_rows)
    haulage_by_pol_zone: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in haulage_rows:
        haulage_by_pol_zone[(row["join_pol"], row["zone_key"])].append(row)

    offers: list[RateOffer] = []
    charges: list[RateChargeLine] = []
    for rate in iter_rate_rows(workbook, rules, card.currency_default):
        matching_haulage = haulage_by_pol_zone.get((rate["join_pol"], rate["zone_key"]), [])
        for haulage in matching_haulage:
            collection = collection_label(haulage, ambiguous_cities)
            for pod, final_destination in expand_destinations(rate["pod"], rate["final_destination"]):
                offer = RateOffer(
                    rate_card_id=card.id,
                    offer_reference=rate["tier"],
                    commodity=card.commodity,
                    origin=collection,
                    place_of_receipt=collection,
                    pol=rate["pol"],
                    pod=pod,
                    final_destination=final_destination,
                    zone=rate["zone_label"],
                    equipment_type=rate["equipment_type"],
                    service_mode="SD / CY",
                    base_amount=rate["amount"],
                    base_currency=rate["currency"],
                    all_in_flag=False,
                    routing_note=rate["freetime"],
                    valid_from=rate["valid_from"] or card.valid_from,
                    valid_to=rate["valid_to"] or card.valid_to,
                    raw_sheet_name=rate["sheet_name"],
                    raw_row_reference=f"{rate['rate_row_reference']} + {haulage['haulage_row_reference']}",
                    raw_row_json={
                        "pricing_tier": rate["tier"],
                        "city": haulage["city"],
                        "area": haulage["area"],
                        "county": haulage["county"],
                        "haulage_pol_raw": haulage["pol"],
                        "rate_pol": rate["pol"],
                        "zone": rate["zone_label"],
                        "pod_raw": rate["pod"],
                        "final_destination_raw": rate["final_destination"],
                        "all_in_rate": rate["amount"],
                        "doc_raw": rate["doc_raw"],
                        "freetime": rate["freetime"],
                        "rate_row_reference": rate["rate_row_reference"],
                        "haulage_row_reference": haulage["haulage_row_reference"],
                    },
                )
                offers.append(offer)
                if rate["doc_amount"] is not None:
                    charges.append(
                        RateChargeLine(
                            rate_offer_id=offer.id,
                            charge_name="Export Booking Documentation Fee",
                            charge_type="origin",
                            basis="per B/L",
                            amount=rate["doc_amount"],
                            currency=rate["currency"],
                            included_flag=False,
                            source_label="DOC (per B/L)",
                            raw_value=rate["doc_raw"],
                        )
                    )

    notes = extract_terms(workbook, template, card.id)
    return card, offers, charges, notes


def extract_tier_rate_tables(path: Path, template: ParserTemplate) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    currency = template.defaults.get("currency_default", "GBP")
    tables: dict[str, list[dict[str, Any]]] = {}
    for rate in iter_rate_rows(workbook, template.inline_haulage_rules, currency):
        tables.setdefault(rate["tier"], []).append(
            {
                "zone": rate["zone_label"],
                "pol": rate["pol"],
                "pod": rate["pod"],
                "final_destination": rate["final_destination"],
                "equipment_type": rate["equipment_type"],
                "amount": rate["amount"],
                "currency": rate["currency"],
                "documentation": rate["doc_raw"],
                "freetime": rate["freetime"],
                "valid_from": rate["valid_from"].isoformat() if rate["valid_from"] else None,
                "valid_to": rate["valid_to"].isoformat() if rate["valid_to"] else None,
                "source_reference": rate["rate_row_reference"],
            }
        )
    return tables


def load_haulage_rows(workbook, rules: dict[str, Any]) -> list[dict[str, str]]:
    sheet_name = rules.get("haulage_sheet", "Haulage Zones")
    sheet = workbook[sheet_name]
    aliases = {normalize_key(left): normalize_key(right) for left, right in rules.get("pol_aliases", {}).items()}
    rows: list[dict[str, str]] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        city = normalize_text(value_at(row, 0))
        pol = normalize_text(value_at(row, 3))
        zone_key = normalize_zone(value_at(row, 4))
        if not city or not pol or not zone_key:
            continue
        raw_pol_key = normalize_key(pol)
        rows.append(
            {
                "city": city,
                "area": normalize_text(value_at(row, 1)),
                "county": normalize_text(value_at(row, 2)),
                "pol": pol,
                "join_pol": aliases.get(raw_pol_key, raw_pol_key),
                "zone_key": zone_key,
                "haulage_row_reference": f"{sheet_name}!R{row_number}",
            }
        )
    return rows


def iter_rate_rows(workbook, rules: dict[str, Any], default_currency: str) -> Iterator[dict[str, Any]]:
    rate_sheets = rules.get(
        "rate_sheets",
        {"REUDAN-SPECIAL": "SPECIAL", "REUDAN-TARRIFF": "TARIFF"},
    )
    for sheet_name, tier in rate_sheets.items():
        sheet = workbook[sheet_name]
        for row_number, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            zone_key = normalize_zone(value_at(row, 0))
            pol = normalize_text(value_at(row, 1))
            pod = normalize_text(value_at(row, 2))
            amount, _ = parse_amount(value_at(row, 5))
            equipment_type, _ = normalize_equipment(
                normalize_text(value_at(row, 4)),
                {"forty_default": template_value(rules, "forty_default", "40HC")},
            )
            if not zone_key or not pol or not pod or not equipment_type or amount is None:
                continue
            doc_raw = normalize_text(value_at(row, 6))
            doc_amount, _ = parse_amount(doc_raw)
            yield {
                "sheet_name": sheet_name,
                "tier": str(tier).upper(),
                "zone_key": zone_key,
                "zone_label": f"ZONE {zone_key}",
                "join_pol": normalize_key(pol),
                "pol": pol,
                "pod": pod,
                "final_destination": clean_nullable(normalize_text(value_at(row, 3))),
                "equipment_type": equipment_type,
                "amount": amount,
                "currency": default_currency,
                "doc_raw": doc_raw,
                "doc_amount": doc_amount,
                "freetime": normalize_text(value_at(row, 7)),
                "valid_from": parse_date_value(value_at(row, 8)),
                "valid_to": parse_date_value(value_at(row, 9)),
                "rate_row_reference": f"{sheet_name}!R{row_number}",
            }


def expand_destinations(pod: str, final_destination: str | None) -> list[tuple[str, str | None]]:
    if final_destination:
        return [(pod, final_destination)]
    destinations = [normalize_text(value) for value in pod.split("/") if normalize_text(value)]
    return [(destination, None) for destination in destinations] or [(pod, None)]


def find_ambiguous_cities(rows: list[dict[str, str]]) -> set[str]:
    locations: dict[str, set[tuple[str, str]]] = defaultdict(set)
    for row in rows:
        locations[normalize_key(row["city"])].add((normalize_key(row["area"]), normalize_key(row["county"])))
    return {city for city, variants in locations.items() if len(variants) > 1}


def collection_label(row: dict[str, str], ambiguous_cities: set[str]) -> str:
    if normalize_key(row["city"]) not in ambiguous_cities:
        return row["city"]
    qualifier = row["county"] or row["area"]
    return f"{row['city']}, {qualifier}" if qualifier else row["city"]


def extract_cover_metadata(workbook) -> dict[str, str]:
    if "Cover page" not in workbook.sheetnames:
        return {}
    metadata: dict[str, str] = {}
    for row in workbook["Cover page"].iter_rows(values_only=True):
        values = [normalize_text(value) for value in row if normalize_text(value)]
        if len(values) >= 2:
            metadata[values[0]] = values[-1]
    return metadata


def extract_terms(workbook, template: ParserTemplate, rate_card_id: str) -> list[RateNote]:
    notes: list[RateNote] = []
    keywords = [str(value).lower() for value in template.note_extraction.get("keywords", [])]
    max_rows = int(template.note_extraction.get("scan_top_rows", 60))
    for sheet_name in workbook.sheetnames:
        if not any(token.upper() in sheet_name.upper() for token in template.note_extraction.get("scan_sheets", [])):
            continue
        sheet = workbook[sheet_name]
        for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(max_rows, sheet.max_row), values_only=True), start=1):
            text = " ".join(normalize_text(value) for value in row if normalize_text(value))
            if text and (not keywords or any(keyword in text.lower() for keyword in keywords)):
                notes.append(
                    RateNote(
                        rate_card_id=rate_card_id,
                        note_type="commercial",
                        note_text=text,
                        source_reference=f"{sheet_name}!R{row_number}",
                    )
                )
    return notes


def normalize_zone(value: object) -> str:
    text = normalize_text(value).upper().replace("ZONE", "").strip()
    if not text:
        return ""
    try:
        return str(int(float(text)))
    except ValueError:
        return text


def parse_compact_date_range(value: object) -> tuple[Any, Any]:
    parsed = parse_date_range(value)
    if parsed[0] and parsed[1]:
        return parsed
    text = normalize_text(value)
    if "-" not in text:
        return parsed
    start, end = text.split("-", 1)
    return parse_date_value(start), parse_date_value(end)


def normalize_key(value: object) -> str:
    return " ".join(normalize_text(value).upper().split())


def clean_nullable(value: str) -> str | None:
    return None if not value or value.upper() in {"N/A", "NA", "N.A."} else value


def value_at(row: tuple[object, ...], index: int) -> object | None:
    return row[index] if index < len(row) else None


def template_value(rules: dict[str, Any], key: str, default: str) -> str:
    value = rules.get(key)
    return str(value) if value is not None else default
