from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from rate_ingest.models import ParserTemplate, RateCard, RateChargeLine, RateImport, RateNote, RateOffer
from rate_ingest.normalize import normalize_text, parse_amount, parse_date_value


def parse_workbook(
    path: Path,
    template: ParserTemplate,
    rate_import: RateImport,
) -> tuple[RateCard, list[RateOffer], list[RateChargeLine], list[RateNote]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    rules = template.hapag_matrix_rules
    sheet_name = rules.get("sheet_name", workbook.sheetnames[0])
    sheet = workbook[sheet_name]

    destination_row = int(rules.get("destination_row", 1))
    routing_row = int(rules.get("routing_row", 2))
    data_start_row = int(rules.get("data_start_row", 4))
    collection_column = int(rules.get("collection_column", 3))
    pol_column = int(rules.get("pol_column", 2))
    destination_start_column = int(rules.get("destination_start_column", 4))
    destination_end_column = int(rules.get("destination_end_column", 10))
    terms_column = int(rules.get("terms_column", 11))

    terms = extract_terms(sheet, terms_column)
    valid_from, valid_to = extract_validity(terms)
    contract_reference = extract_contract_reference(terms)
    currency = template.defaults.get("currency_default", "USD")
    card = RateCard(
        rate_import_id=rate_import.id,
        provider_name=template.provider_name,
        carrier_name=template.defaults.get("carrier_name", template.provider_name),
        document_type=template.document_type,
        commodity=template.defaults.get("commodity"),
        currency_default=currency,
        valid_from=valid_from,
        valid_to=valid_to,
        all_in_flag=False,
        notes_summary="Hapag-Lloyd door-to-quay matrix; live position and conditional emergency fuel are additional.",
    )

    destinations = {
        column: {
            "pod": normalize_text(sheet.cell(destination_row, column).value),
            "routing": normalize_text(sheet.cell(routing_row, column).value),
        }
        for column in range(destination_start_column, destination_end_column + 1)
    }
    emergency_destinations = {
        normalize_key(value)
        for value in rules.get("emergency_fuel_destinations", ["Binh Duong", "Lat Krabang"])
    }
    live_position_amount = float(rules.get("live_position_amount", 15))
    emergency_fuel_amount = float(rules.get("emergency_fuel_amount", 20))

    offers: list[RateOffer] = []
    charges: list[RateChargeLine] = []
    for row_number in range(data_start_row, sheet.max_row + 1):
        collection_raw = normalize_text(sheet.cell(row_number, collection_column).value)
        pol = normalize_text(sheet.cell(row_number, pol_column).value)
        if not collection_raw or not pol:
            continue
        collection = clean_collection(collection_raw)
        collection_source_code = extract_collection_source_code(collection_raw)
        for column, destination in destinations.items():
            pod = destination["pod"]
            amount, trailing_note = parse_amount(sheet.cell(row_number, column).value)
            if not pod or amount is None:
                continue
            routing_parts = [part for part in [destination["routing"], trailing_note] if part]
            offer = RateOffer(
                rate_card_id=card.id,
                offer_reference=contract_reference,
                commodity=card.commodity,
                origin=collection,
                place_of_receipt=collection,
                pol=pol,
                pod=pod,
                final_destination=pod,
                equipment_type=template.defaults.get("equipment_type", "40HC"),
                service_mode="SD / CY",
                base_amount=amount,
                base_currency=currency,
                all_in_flag=False,
                routing_note=" | ".join(routing_parts) or None,
                valid_from=valid_from,
                valid_to=valid_to,
                raw_sheet_name=sheet_name,
                raw_row_reference=f"{sheet_name}!R{row_number}C{column}",
                raw_row_json={
                    "collection_raw": collection_raw,
                    "collection_source_code": collection_source_code,
                    "preferred_pol": pol,
                    "destination": pod,
                    "applicable_routing": destination["routing"],
                    "matrix_value": normalize_text(sheet.cell(row_number, column).value),
                },
            )
            offers.append(offer)
            charges.append(
                RateChargeLine(
                    rate_offer_id=offer.id,
                    charge_name="Live Position",
                    charge_type="origin",
                    basis="per_container",
                    amount=live_position_amount,
                    currency=currency,
                    included_flag=False,
                    source_label="Live Position USD 15/container",
                    raw_value=f"{currency} {live_position_amount:g}/container",
                )
            )
            if normalize_key(pod) in emergency_destinations:
                charges.append(
                    RateChargeLine(
                        rate_offer_id=offer.id,
                        charge_name="Emergency Fuel Destination",
                        charge_type="destination",
                        basis="per_container",
                        amount=emergency_fuel_amount,
                        currency=currency,
                        included_flag=False,
                        source_label="Emergency Fuel Destination USD 20/container",
                        raw_value=f"{currency} {emergency_fuel_amount:g}/container",
                    )
                )

    notes = [
        RateNote(
            rate_card_id=card.id,
            note_type="commercial",
            note_text=text,
            source_reference=f"{sheet_name}!R{row_number}C{terms_column}",
        )
        for row_number, text in terms
    ]
    return card, offers, charges, notes


def extract_terms(sheet, terms_column: int) -> list[tuple[int, str]]:
    return [
        (row_number, text)
        for row_number in range(1, sheet.max_row + 1)
        if (text := normalize_text(sheet.cell(row_number, terms_column).value))
    ]


def extract_validity(terms: list[tuple[int, str]]):
    for _, text in terms:
        match = re.search(r"validity\s+(.+?)\s+(?:to|until)\s+(.+)$", text, re.IGNORECASE)
        if match:
            return parse_date_value(match.group(1)), parse_date_value(match.group(2))
    return None, None


def extract_contract_reference(terms: list[tuple[int, str]]) -> str | None:
    for _, text in terms:
        match = re.search(r"contract\s+number\s+(.+)$", text, re.IGNORECASE)
        if match:
            return normalize_text(match.group(1))
    return None


def clean_collection(raw: str) -> str:
    location = raw.split("/", 1)[-1].strip()
    return location.title()


def extract_collection_source_code(raw: str) -> str | None:
    match = re.match(r"^\s*([A-Za-z]{2}[A-Za-z0-9]{3})\s*/", raw)
    return match.group(1).upper() if match else None


def normalize_key(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", " ", normalize_text(value).upper()).strip()
