from __future__ import annotations

from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from rate_ingest.models import ParserTemplate, RateCard, RateChargeLine, RateImport, RateNote, RateOffer
from rate_ingest.normalize import normalize_equipment, normalize_text, parse_amount, parse_date_value


def parse_workbook(
    path: Path, template: ParserTemplate, rate_import: RateImport
) -> tuple[RateCard, list[RateOffer], list[RateChargeLine], list[RateNote]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    rules = template.site_to_site_rules
    quote_sheet_name = select_sheet_name(workbook.sheetnames, rules.get("quote_sheet_name_contains", "AFLS Quote"))
    quote_sheet = workbook[quote_sheet_name]
    header_row = int(rules.get("header_row", 7))
    headers = [normalize_text(quote_sheet.cell(header_row, column).value) for column in range(1, quote_sheet.max_column + 1)]
    field_indexes = {field: find_header_index(headers, label) for field, label in template.field_map.items()}
    rate_basis_index = field_indexes.get("rate_basis")
    amount_columns = detect_amount_columns(headers, rate_basis_index)
    if not amount_columns:
        raise ValueError("No equipment amount columns were found in the AFLS Quote sheet.")

    metadata = extract_top_metadata(quote_sheet, header_row)
    charge_name_map = load_charge_name_map(workbook, rules.get("abbreviation_sheet_name_contains", "Abbreviation"))
    card = RateCard(
        rate_import_id=rate_import.id,
        provider_name=template.provider_name,
        carrier_name=template.defaults.get("carrier_name", template.provider_name),
        document_type=template.document_type,
        commodity=None,
        currency_default=template.defaults.get("currency_default"),
        valid_from=None,
        valid_to=None,
        all_in_flag=False,
        notes_summary=None,
    )

    offers_by_key: "OrderedDict[tuple[Any, ...], RateOffer]" = OrderedDict()
    row_bounds: dict[tuple[Any, ...], list[int]] = {}
    charges: list[RateChargeLine] = []
    notes: list[RateNote] = []

    for row_number, row in enumerate(
        quote_sheet.iter_rows(min_row=header_row + 1, max_row=quote_sheet.max_row, values_only=True),
        start=header_row + 1,
    ):
        receipt = clean_nullable(normalize_text(value_at(row, field_indexes.get("place_of_receipt"))))
        delivery = clean_nullable(normalize_text(value_at(row, field_indexes.get("final_destination"))))
        charge_code = clean_nullable(normalize_text(value_at(row, field_indexes.get("charge_code"))))
        if not receipt or not delivery or not charge_code:
            continue

        pol = clean_nullable(normalize_text(value_at(row, field_indexes.get("pol"))))
        pod = clean_nullable(normalize_text(value_at(row, field_indexes.get("pod"))))
        service_mode = clean_nullable(normalize_text(value_at(row, field_indexes.get("service_mode"))))
        commodity = clean_nullable(normalize_text(value_at(row, field_indexes.get("commodity"))))
        transit_time_days = parse_transit_time(value_at(row, field_indexes.get("transit_time")))
        valid_from = parse_date_cell(value_at(row, field_indexes.get("valid_from")))
        valid_to = parse_date_cell(value_at(row, field_indexes.get("valid_to")))
        rate_basis = clean_nullable(normalize_text(value_at(row, rate_basis_index)))

        for equipment_header, amount_index in amount_columns:
            raw_amount = value_at(row, amount_index)
            amount, tail = parse_amount(raw_amount)
            if amount is None:
                continue

            equipment_type, _ = normalize_equipment(strip_equipment_header(equipment_header), template.defaults)
            charge_currency = normalize_currency_code(tail) or template.defaults.get("currency_default")
            offer_key = (
                receipt,
                pol,
                pod,
                delivery,
                service_mode,
                commodity,
                transit_time_days,
                valid_from,
                valid_to,
                equipment_type,
            )
            offer = offers_by_key.get(offer_key)
            if offer is None:
                offer = RateOffer(
                    rate_card_id=card.id,
                    offer_reference=metadata.get("Quote Number"),
                    origin=receipt,
                    place_of_receipt=receipt,
                    pol=pol,
                    pod=pod,
                    final_destination=delivery,
                    equipment_type=equipment_type,
                    service_mode=service_mode,
                    transit_time_days=transit_time_days,
                    base_amount=None,
                    base_currency=template.defaults.get("currency_default"),
                    all_in_flag=False,
                    routing_note=None,
                    valid_from=valid_from,
                    valid_to=valid_to,
                    raw_sheet_name=quote_sheet_name,
                    raw_row_reference=f"{quote_sheet_name}!R{row_number}",
                    raw_row_json={
                        "quote_number": metadata.get("Quote Number"),
                        "customer_name": metadata.get("Customer Name"),
                        "carrier": metadata.get("Carrier"),
                        "last_acceptance_date": metadata.get("Last Acceptance Date"),
                    },
                )
                offers_by_key[offer_key] = offer
                row_bounds[offer_key] = [row_number, row_number]
            else:
                row_bounds[offer_key][1] = row_number

            charge_name = charge_name_map.get(charge_code, charge_code)
            if charge_code == "BAS":
                offer.base_amount = amount
                offer.base_currency = charge_currency or offer.base_currency

            charges.append(
                RateChargeLine(
                    rate_offer_id=offer.id,
                    charge_name=charge_name,
                    charge_type=infer_charge_type(charge_name),
                    basis=normalize_basis(rate_basis),
                    amount=amount,
                    currency=charge_currency,
                    included_flag=False if amount not in {0, None} else "unknown",
                    source_label=charge_code,
                    raw_value=normalize_text(raw_amount),
                )
            )

            if card.commodity is None and commodity:
                card.commodity = commodity
            if card.valid_from is None or (valid_from and valid_from < card.valid_from):
                card.valid_from = valid_from
            if card.valid_to is None or (valid_to and valid_to > card.valid_to):
                card.valid_to = valid_to

    offers = list(offers_by_key.values())
    for offer_key, offer in offers_by_key.items():
        start_row, end_row = row_bounds[offer_key]
        offer.raw_row_reference = f"{quote_sheet_name}!R{start_row}:R{end_row}"

    if metadata.get("Quote Number"):
        notes.append(
            RateNote(
                rate_card_id=card.id,
                note_type="commercial",
                note_text=f"Quote Number {metadata['Quote Number']}",
                source_reference=f"{quote_sheet_name}!R1",
            )
        )
    if metadata.get("Last Acceptance Date"):
        notes.append(
            RateNote(
                rate_card_id=card.id,
                note_type="commercial",
                note_text=f"Last Acceptance Date {metadata['Last Acceptance Date']}",
                source_reference=f"{quote_sheet_name}!R3",
            )
        )
    if notes:
        card.notes_summary = notes[0].note_text[:240]
    return card, offers, charges, notes


def select_sheet_name(sheet_names: list[str], contains_token: str) -> str:
    needle = normalize_text(contains_token).upper()
    for sheet_name in sheet_names:
        if needle in normalize_text(sheet_name).upper():
            return sheet_name
    return sheet_names[0]


def find_header_index(headers: list[str], label: str) -> int | None:
    needle = normalize_text(label).upper()
    for index, header in enumerate(headers):
        if needle and needle == normalize_text(header).upper():
            return index
    return None


def detect_amount_columns(headers: list[str], rate_basis_index: int | None) -> list[tuple[str, int]]:
    if rate_basis_index is None:
        return []
    amount_columns: list[tuple[str, int]] = []
    for index in range(rate_basis_index + 1, len(headers)):
        header = normalize_text(headers[index])
        if not header:
            continue
        amount_columns.append((header, index))
    return amount_columns


def load_charge_name_map(workbook, contains_token: str) -> dict[str, str]:
    sheet_name = select_sheet_name(workbook.sheetnames, contains_token)
    sheet = workbook[sheet_name]
    charge_names: dict[str, str] = {}
    for row in sheet.iter_rows(values_only=True):
        code = normalize_text(row[0] if len(row) > 0 else None)
        description = normalize_text(row[1] if len(row) > 1 else None)
        if code == "Charge_Code":
            continue
        if code == "Container Size-Type":
            break
        if code and description:
            charge_names[code] = description
    for row in sheet.iter_rows(values_only=True):
        code = normalize_text(row[0] if len(row) > 0 else None)
        description = normalize_text(row[1] if len(row) > 1 else None)
        if code and description and code not in charge_names and code not in {"Charge_Code", "Container Size-Type"}:
            charge_names[code] = description
    return charge_names


def extract_top_metadata(sheet, header_row: int) -> dict[str, str]:
    metadata: dict[str, str] = {}
    for row_number in range(1, header_row):
        key = normalize_text(sheet.cell(row_number, 1).value)
        value = sheet.cell(row_number, 2).value
        if key and value not in (None, ""):
            metadata[key] = stringify_date(value)
    return metadata


def value_at(row: tuple[object, ...], index: int | None) -> object | None:
    if index is None or index >= len(row):
        return None
    return row[index]


def strip_equipment_header(value: str) -> str:
    return normalize_text(value).replace(":", "").strip()


def normalize_currency_code(value: str | None) -> str | None:
    text = normalize_text(value).upper()
    if not text:
        return None
    letters = "".join(ch for ch in text if ch.isalpha())
    return letters or None


def normalize_basis(value: str | None) -> str | None:
    text = normalize_text(value).upper()
    if not text:
        return None
    mapping = {
        "PER_CONTAINER": "Container",
        "PER_DOC": "Per document",
        "PER_DOCUMENT": "Per document",
    }
    return mapping.get(text, text.replace("_", " ").title())


def infer_charge_type(charge_name: str) -> str | None:
    text = normalize_text(charge_name).lower()
    if any(token in text for token in ("origin", "export", "inland haulage export")):
        return "origin"
    if any(token in text for token in ("destination", "import", "inland haulage import")):
        return "destination"
    if any(token in text for token in ("ocean freight", "bunker", "peak season", "emission", "risk surcharge", "contingency", "freetime extension", "free out")):
        return "freight"
    return None


def parse_transit_time(value: object) -> int | None:
    text = normalize_text(value)
    if not text:
        return None
    digits = "".join(ch for ch in text if ch.isdigit())
    return int(digits) if digits else None


def parse_date_cell(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return parse_date_value(value)


def stringify_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return normalize_text(value)


def clean_nullable(value: str | None) -> str | None:
    if not value:
        return None
    if value.upper() in {"N/A", "NA", "NONE"}:
        return None
    return value
