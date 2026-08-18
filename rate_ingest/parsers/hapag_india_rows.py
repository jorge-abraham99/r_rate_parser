from __future__ import annotations

from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from rate_ingest.models import (
    ParserTemplate,
    RateCard,
    RateChargeLine,
    RateImport,
    RateNote,
    RateOffer,
)
from rate_ingest.normalize import normalize_equipment, normalize_text, parse_amount, parse_date_value


def parse_workbook(
    path: Path,
    template: ParserTemplate,
    rate_import: RateImport,
) -> tuple[RateCard, list[RateOffer], list[RateChargeLine], list[RateNote]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    rules = template.hapag_india_rules
    sheet_name = select_sheet_name(workbook.sheetnames, rules.get("detail_sheet_name", "Detail"))
    sheet = workbook[sheet_name]

    header_row = int(rules.get("header_row", 18))
    data_start_row = int(rules.get("data_start_row", header_row + 1))
    columns = {
        key: int(value) - 1
        for key, value in rules.get("columns", {}).items()
    }
    charge_codes = {
        normalize_code(value)
        for value in rules.get("charge_codes", [])
    }
    total_charge_codes = [
        normalize_code(value)
        for value in rules.get("total_charge_codes", ["Lumpsum", "LPC", "EOD"])
    ]

    valid_from = parse_sheet_date(
        sheet.cell(int(rules.get("validity_row", 3)), int(rules.get("valid_from_column", 18))).value
    )
    valid_to = parse_sheet_date(
        sheet.cell(int(rules.get("validity_row", 3)), int(rules.get("valid_to_column", 19))).value
    )
    quote_number = normalize_text(
        sheet.cell(int(rules.get("quote_number_row", 10)), int(rules.get("quote_number_column", 18))).value
    ) or None

    card = RateCard(
        rate_import_id=rate_import.id,
        provider_name=template.provider_name,
        carrier_name=template.defaults.get("carrier_name", template.provider_name),
        document_type=template.document_type,
        commodity=template.defaults.get("commodity"),
        currency_default=template.defaults.get("currency_default", "USD"),
        valid_from=valid_from,
        valid_to=valid_to,
        all_in_flag=False,
        notes_summary=(f"Hapag-Lloyd India quote {quote_number}" if quote_number else None),
    )

    offers_by_key: "OrderedDict[tuple[Any, ...], RateOffer]" = OrderedDict()
    row_bounds: dict[tuple[Any, ...], list[int]] = {}
    charges: list[RateChargeLine] = []

    for row_number, row in enumerate(
        sheet.iter_rows(min_row=data_start_row, max_row=sheet.max_row, values_only=True),
        start=data_start_row,
    ):
        charge_code = normalize_code(value_at(row, columns.get("charge_code")))
        if not charge_code or charge_code not in charge_codes:
            continue

        origin_raw = normalize_text(value_at(row, columns.get("origin")))
        pol_raw = normalize_text(value_at(row, columns.get("pol")))
        pod_raw = normalize_text(value_at(row, columns.get("pod")))
        equipment_raw = normalize_text(value_at(row, columns.get("equipment")))
        if not origin_raw or not pol_raw or not pod_raw or not equipment_raw:
            continue

        equipment_type, _ = normalize_equipment(equipment_raw, template.defaults)
        if not equipment_type:
            continue
        amount, trailing_note = parse_amount(value_at(row, columns.get("amount")))
        if amount is None:
            continue

        origin = clean_place(origin_raw)
        pol = clean_place(pol_raw)
        pod = clean_place(pod_raw)
        via = clean_place(value_at(row, columns.get("via"))) or None
        commodity = normalize_text(value_at(row, columns.get("commodity"))) or card.commodity
        currency = normalize_text(value_at(row, columns.get("currency"))).upper() or card.currency_default
        service_mode = service_mode_for(
            value_at(row, columns.get("export_haulage")),
            value_at(row, columns.get("import_haulage")),
        )
        transit_time = parse_integer(value_at(row, columns.get("transit_time")))
        service_id = normalize_text(value_at(row, columns.get("service_id")))
        offer_key = (
            service_id,
            origin,
            pol,
            pod,
            equipment_type,
            service_mode,
            commodity,
            transit_time,
        )

        offer = offers_by_key.get(offer_key)
        if offer is None:
            offer = RateOffer(
                rate_card_id=card.id,
                offer_reference=quote_number,
                commodity=commodity,
                origin=origin,
                place_of_receipt=origin,
                pol=pol,
                pod=pod,
                final_destination=pod,
                equipment_type=equipment_type,
                service_mode=service_mode,
                transit_time_days=transit_time,
                base_amount=None,
                base_currency=currency,
                all_in_flag=False,
                routing_note=via,
                valid_from=valid_from,
                valid_to=valid_to,
                raw_sheet_name=sheet_name,
                raw_row_reference=f"{sheet_name}!R{row_number}",
                raw_row_json={
                    "service_id": service_id,
                    "total_charge_codes": total_charge_codes,
                    "quote_number": quote_number,
                    "origin_raw": origin_raw,
                    "port_of_loading_raw": pol_raw,
                    "port_of_discharge_raw": pod_raw,
                },
            )
            offers_by_key[offer_key] = offer
            row_bounds[offer_key] = [row_number, row_number]
        else:
            row_bounds[offer_key][1] = row_number

        if charge_code == "LUMPSUM":
            offer.base_amount = amount
            offer.base_currency = currency

        charges.append(
            RateChargeLine(
                rate_offer_id=offer.id,
                charge_name=display_charge_name(charge_code),
                charge_type=charge_type_for(value_at(row, columns.get("charge_type")), charge_code),
                basis=normalize_basis(value_at(row, columns.get("unit"))),
                amount=amount,
                currency=currency,
                included_flag=False,
                source_label=charge_code,
                raw_value=build_raw_value(amount, currency, value_at(row, columns.get("unit")), trailing_note),
            )
        )

        if card.commodity is None and commodity:
            card.commodity = commodity

    offers = list(offers_by_key.values())
    for offer_key, offer in offers_by_key.items():
        start_row, end_row = row_bounds[offer_key]
        offer.raw_row_reference = f"{sheet_name}!R{start_row}:R{end_row}"

    notes = build_notes(card.id, sheet_name, quote_number, valid_from, valid_to, workbook, rules)
    if notes:
        card.notes_summary = notes[0].note_text[:240]
    return card, offers, charges, notes


def select_sheet_name(sheet_names: list[str], expected_name: str) -> str:
    needle = normalize_text(expected_name).upper()
    for sheet_name in sheet_names:
        if needle in normalize_text(sheet_name).upper():
            return sheet_name
    return sheet_names[0]


def value_at(row: tuple[object, ...], index: int | None) -> object | None:
    if index is None or index >= len(row):
        return None
    return row[index]


def normalize_code(value: object) -> str:
    return normalize_text(value).upper()


def parse_sheet_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return parse_date_value(value)


def clean_place(value: object) -> str:
    text = normalize_text(value)
    return text.title() if text else ""


def parse_integer(value: object) -> int | None:
    text = normalize_text(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        match = re.search(r"\d+", text)
        return int(match.group(0)) if match else None


def service_mode_for(export_haulage: object, import_haulage: object) -> str | None:
    export_mode = normalize_text(export_haulage).upper()
    import_mode = normalize_text(import_haulage).upper()
    if export_mode == "DOOR" and import_mode == "PORT":
        return "SD / CY"
    if export_mode and import_mode:
        return f"{export_mode.title()} / {import_mode.title()}"
    return None


def charge_type_for(charge_type: object, charge_code: str) -> str | None:
    source_type = normalize_text(charge_type).lower()
    if charge_code == "LUMPSUM":
        return "base"
    if "export" in source_type:
        return "origin"
    if "import" in source_type:
        return "destination"
    if "freight" in source_type:
        return "freight"
    return None


def normalize_basis(value: object) -> str | None:
    text = normalize_text(value).upper()
    if not text:
        return None
    return {
        "CTR": "Container",
        "CONTAINER": "Container",
        "BIL": "Per document",
        "PER BILL OF LADING": "Per document",
    }.get(text, text.replace("_", " ").title())


def display_charge_name(charge_code: str) -> str:
    return "Lumpsum" if charge_code == "LUMPSUM" else charge_code


def build_raw_value(amount: float, currency: str, unit: object, trailing_note: str | None) -> str:
    parts = [f"{currency} {amount:g}", normalize_text(unit)]
    if trailing_note:
        parts.append(trailing_note)
    return " ".join(part for part in parts if part)


def build_notes(
    card_id: str,
    sheet_name: str,
    quote_number: str | None,
    valid_from,
    valid_to,
    workbook,
    rules: dict[str, Any],
) -> list[RateNote]:
    notes: list[RateNote] = []
    if quote_number:
        notes.append(
            RateNote(
                rate_card_id=card_id,
                note_type="commercial",
                note_text=f"Quotation number {quote_number}",
                source_reference=f"{sheet_name}!R{int(rules.get('quote_number_row', 10))}C{int(rules.get('quote_number_column', 18))}",
            )
        )
    if valid_from or valid_to:
        notes.append(
            RateNote(
                rate_card_id=card_id,
                note_type="commercial",
                note_text=f"Validity {valid_from or '—'} to {valid_to or '—'}",
                source_reference=f"{sheet_name}!R{int(rules.get('validity_row', 3))}",
            )
        )
    terms_name = select_sheet_name(workbook.sheetnames, rules.get("terms_sheet_name", "Terms"))
    terms_sheet = workbook[terms_name]
    for row_number, row in enumerate(terms_sheet.iter_rows(values_only=True), start=1):
        text = next((normalize_text(value) for value in row if normalize_text(value)), "")
        if text and len(text) >= 40 and any(token in text.lower() for token in ("charge", "valid", "booking", "freight")):
            notes.append(
                RateNote(
                    rate_card_id=card_id,
                    note_type="commercial",
                    note_text=text,
                    source_reference=f"{terms_name}!R{row_number}",
                )
            )
            if len(notes) >= 12:
                break
    return notes
