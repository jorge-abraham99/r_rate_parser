from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from rate_ingest.models import ParserTemplate, RateCard, RateChargeLine, RateImport, RateNote, RateOffer
from rate_ingest.normalize import normalize_text, parse_amount, parse_date_value


def parse_workbook(
    path: Path, template: ParserTemplate, rate_import: RateImport
) -> tuple[RateCard, list[RateOffer], list[RateChargeLine], list[RateNote]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet_name = select_sheet(workbook.sheetnames, template)
    sheet = workbook[sheet_name]
    rules = template.matrix_rules
    header_row = int(rules.get("header_row", 2))
    data_start_row = int(rules.get("data_start_row", header_row + 2))
    origin_column = int(rules.get("origin_column", 1))
    destination_start_column = int(rules.get("destination_start_column", 2))
    destination_end_column = int(rules.get("destination_end_column", sheet.max_column))

    top_left_note = normalize_text(sheet.cell(header_row, origin_column).value)
    top_title = normalize_text(sheet.cell(1, 1).value)
    valid_from, valid_to = parse_matrix_validity(top_left_note)
    pol_text = extract_pol(top_left_note)

    card = RateCard(
        rate_import_id=rate_import.id,
        provider_name=template.provider_name,
        carrier_name=template.defaults.get("carrier_name", template.provider_name),
        document_type=template.document_type,
        commodity=template.defaults.get("commodity"),
        currency_default=template.defaults.get("currency_default"),
        valid_from=valid_from,
        valid_to=valid_to,
        all_in_flag=template.defaults.get("all_in_flag", False),
        notes_summary=top_left_note or top_title or None,
    )

    destination_headers = {}
    for column in range(destination_start_column, destination_end_column + 1):
        parsed = parse_destination_header(normalize_text(sheet.cell(header_row, column).value))
        destination_headers[column] = parsed

    offers: list[RateOffer] = []
    notes: list[RateNote] = []
    if top_title:
        notes.append(
            RateNote(
                rate_card_id=card.id,
                note_type="general",
                note_text=top_title,
                source_reference=f"{sheet_name}!R1",
            )
        )
    if top_left_note:
        notes.append(
            RateNote(
                rate_card_id=card.id,
                note_type="commercial",
                note_text=top_left_note,
                source_reference=f"{sheet_name}!R{header_row}C{origin_column}",
            )
        )

    for row_number in range(data_start_row, sheet.max_row + 1):
        origin = normalize_text(sheet.cell(row_number, origin_column).value)
        if not origin:
            continue
        for column in range(destination_start_column, destination_end_column + 1):
            raw_value = sheet.cell(row_number, column).value
            amount, trailing_note = parse_amount(raw_value)
            if amount is None:
                continue
            header = destination_headers[column]
            note_parts = [part for part in [header.get("routing_note"), trailing_note] if part]
            offers.append(
                RateOffer(
                    rate_card_id=card.id,
                    origin=origin,
                    place_of_receipt=origin,
                    pol=pol_text,
                    pod=header.get("to_raw"),
                    final_destination=header.get("to_raw"),
                    equipment_type=template.defaults.get("equipment_type", "UNSPECIFIED"),
                    base_amount=amount,
                    base_currency=template.defaults.get("currency_default"),
                    all_in_flag=template.defaults.get("all_in_flag", False),
                    routing_note=" | ".join(note_parts) or None,
                    valid_from=valid_from,
                    valid_to=valid_to,
                    raw_sheet_name=sheet_name,
                    raw_row_reference=f"{sheet_name}!R{row_number}C{column}",
                    raw_row_json={
                        "origin": origin,
                        "destination_header": header.get("raw_header"),
                        "raw_value": normalize_text(raw_value),
                    },
                )
            )

    return card, offers, [], notes


def select_sheet(sheet_names: list[str], template: ParserTemplate) -> str:
    includes = [token.upper() for token in template.sheet_rules.get("include_sheet_name_contains", [])]
    for sheet_name in sheet_names:
        if not includes or any(token in sheet_name.upper() for token in includes):
            return sheet_name
    return sheet_names[0]


def parse_matrix_validity(text: str):
    match = re.search(
        r"valid from\s+(\d{2}\.\d{2}\.\d{2,4}).*?until\s+(\d{2}\.\d{2}\.\d{2,4})",
        text,
        re.IGNORECASE,
    )
    if not match:
        return None, None
    return parse_date_value(match.group(1)), parse_date_value(match.group(2))


def extract_pol(text: str) -> str | None:
    match = re.search(r"POL\s+(.+?)\s+All rates are subject", text, re.IGNORECASE)
    if not match:
        return None
    return normalize_text(match.group(1))


def parse_destination_header(raw_header: str) -> dict[str, str | None]:
    header = raw_header.strip()
    if not header:
        return {"raw_header": "", "to_raw": None, "routing_note": None}

    cleaned = re.sub(r"\(.*?\)", "", header).strip()
    via_match = re.search(r"\b(via .+)$", cleaned, re.IGNORECASE)
    routing_note = via_match.group(1).strip() if via_match else None
    if via_match:
        cleaned = cleaned[: via_match.start()].strip()

    primary = cleaned.split("/")[0].strip()
    primary = re.sub(r"\b[A-Z]{2,}\d+\b.*$", "", primary).strip()
    primary = primary.rstrip(",")
    return {"raw_header": header, "to_raw": primary or header, "routing_note": routing_note}
