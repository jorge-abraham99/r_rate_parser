from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from rate_ingest.models import ParserTemplate, RateCard, RateChargeLine, RateImport, RateNote, RateOffer
from rate_ingest.normalize import normalize_text, parse_amount, parse_date_range


def parse_workbook(
    path: Path, template: ParserTemplate, rate_import: RateImport
) -> tuple[RateCard, list[RateOffer], list[RateChargeLine], list[RateNote]]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    sheet_name = select_sheet(workbook.sheetnames, template)
    sheet = workbook[sheet_name]
    rules = template.offer_block_rules
    offer_prefix = rules.get("offer_prefix", "Offer")
    offer_start_rows = []
    for row_number in range(1, sheet.max_row + 1):
        first_value = normalize_text(sheet.cell(row_number, 1).value)
        if first_value.startswith(offer_prefix):
            offer_start_rows.append(row_number)

    card = RateCard(
        rate_import_id=rate_import.id,
        provider_name=template.provider_name,
        carrier_name=template.defaults.get("carrier_name", template.provider_name),
        document_type=template.document_type,
        commodity=template.defaults.get("commodity"),
        currency_default=template.defaults.get("currency_default"),
        valid_from=None,
        valid_to=None,
        all_in_flag=template.defaults.get("all_in_flag", False),
        notes_summary=None,
    )

    offers: list[RateOffer] = []
    charges: list[RateChargeLine] = []
    notes: list[RateNote] = []
    offer_start_rows.append(sheet.max_row + 1)

    for index in range(len(offer_start_rows) - 1):
        block_start = offer_start_rows[index]
        block_end = offer_start_rows[index + 1] - 1
        offer, block_charges, block_notes = parse_offer_block(
            sheet_name, sheet, block_start, block_end, card, template
        )
        if offer:
            offers.append(offer)
            charges.extend(block_charges)
            notes.extend(block_notes)
            if card.valid_from is None:
                card.valid_from = offer.valid_from
            if card.valid_to is None:
                card.valid_to = offer.valid_to
            if card.commodity is None and offer.raw_row_json.get("commodity"):
                card.commodity = offer.raw_row_json["commodity"]

    if notes:
        card.notes_summary = notes[0].note_text[:240]
    return card, offers, charges, notes


def select_sheet(sheet_names: list[str], template: ParserTemplate) -> str:
    includes = [token.upper() for token in template.sheet_rules.get("include_sheet_name_contains", [])]
    for sheet_name in sheet_names:
        if not includes or any(token in sheet_name.upper() for token in includes):
            return sheet_name
    return sheet_names[0]


def parse_offer_block(sheet_name, sheet, start_row: int, end_row: int, card: RateCard, template: ParserTemplate):
    offer_reference = normalize_text(sheet.cell(start_row, 1).value)
    metadata: dict[str, str] = {}
    charge_header_row = None

    for row_number in range(start_row + 1, end_row + 1):
        first = normalize_text(sheet.cell(row_number, 1).value)
        second = normalize_text(sheet.cell(row_number, 2).value)
        third = normalize_text(sheet.cell(row_number, 3).value)
        fourth = normalize_text(sheet.cell(row_number, 4).value)
        if first == "Surcharge Name":
            charge_header_row = row_number
            break
        if first and second:
            metadata[first] = second
        if third and fourth:
            metadata[third] = fourth

    valid_from, valid_to = parse_date_range(metadata.get("Rate Validity"))
    base_amount = None
    base_currency = None
    equipment_type = template.defaults.get("equipment_type", "40HDRY")
    block_charges: list[RateChargeLine] = []
    block_notes: list[RateNote] = []

    if charge_header_row:
        equipment_type = normalize_text(sheet.cell(charge_header_row, 4).value) or equipment_type
        for row_number in range(charge_header_row + 1, end_row + 1):
            charge_name = normalize_text(sheet.cell(row_number, 1).value)
            if not charge_name:
                continue
            basis = normalize_text(sheet.cell(row_number, 2).value) or None
            currency = normalize_text(sheet.cell(row_number, 3).value) or None
            amount, _ = parse_amount(sheet.cell(row_number, 4).value)
            charge_type = normalize_text(sheet.cell(row_number, 5).value) or None
            if amount is None:
                continue
            if charge_name.upper().startswith("BASIC OCEAN FREIGHT"):
                base_amount = amount
                base_currency = currency
            block_charges.append(
                RateChargeLine(
                    rate_offer_id="__PENDING__",
                    charge_name=charge_name,
                    charge_type=charge_type.lower() if charge_type else None,
                    basis=basis,
                    amount=amount,
                    currency=currency,
                    included_flag=False if amount not in {0, None} else "unknown",
                    source_label=charge_name,
                    raw_value=f"{charge_name} | {basis or ''} | {currency or ''} | {amount}",
                )
            )

    raw_place_of_receipt = metadata.get("Place of Receipt")
    raw_place_of_delivery = metadata.get("Place of Delivery")
    offer = RateOffer(
        rate_card_id=card.id,
        offer_reference=offer_reference,
        origin=raw_place_of_receipt,
        place_of_receipt=raw_place_of_receipt,
        final_destination=raw_place_of_delivery,
        equipment_type=equipment_type,
        service_mode=metadata.get("Service Mode"),
        transit_time_days=parse_transit_time(metadata.get("Transit Time")),
        base_amount=base_amount,
        base_currency=base_currency or template.defaults.get("currency_default"),
        all_in_flag=False,
        valid_from=valid_from,
        valid_to=valid_to,
        raw_sheet_name=sheet_name,
        raw_row_reference=f"{sheet_name}!R{start_row}:R{end_row}",
        raw_row_json={
            "offer_reference": offer_reference,
            "commodity": metadata.get("Commodity"),
            "mode_of_transport": metadata.get("Mode of Transport"),
            "scheduled_route": metadata.get("Scheduled Route"),
        },
    )
    for charge in block_charges:
        charge.rate_offer_id = offer.id

    route_text = metadata.get("Scheduled Route")
    if route_text:
        block_notes.append(
            RateNote(
                rate_card_id=card.id,
                rate_offer_id=offer.id,
                note_type="routing",
                note_text=route_text,
                source_reference=f"{sheet_name}!R{start_row + 1}",
            )
        )
    return offer, block_charges, block_notes


def parse_transit_time(value: str | None) -> int | None:
    text = normalize_text(value)
    if not text:
        return None
    digits = "".join(ch for ch in text if ch.isdigit())
    return int(digits) if digits else None
