from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from rate_ingest.models import ParserTemplate, RateCard, RateChargeLine, RateNote, RateOffer, RateImport
from rate_ingest.normalize import normalize_equipment, normalize_text, parse_amount, parse_date_range


def parse_workbook(path: Path, template: ParserTemplate, rate_import: RateImport) -> tuple[RateCard, list[RateOffer], list[RateChargeLine], list[RateNote]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    include_tokens = [token.upper() for token in template.sheet_rules.get("include_sheet_name_contains", [])]
    exclude_tokens = [token.upper() for token in template.sheet_rules.get("exclude_sheet_name_contains", [])]

    matched_sheets = []
    for sheet_name in workbook.sheetnames:
        upper = sheet_name.upper()
        if include_tokens and not any(token in upper for token in include_tokens):
            continue
        if any(token in upper for token in exclude_tokens):
            continue
        matched_sheets.append(sheet_name)

    metadata = extract_metadata(workbook)
    valid_from, valid_to = parse_date_range(metadata.get("Validity From/To"))
    card = RateCard(
        rate_import_id=rate_import.id,
        provider_name=template.provider_name,
        carrier_name=template.provider_name,
        document_type=template.document_type,
        commodity=metadata.get("Commodity"),
        currency_default=template.normalizers.get("currency", {}).get("default"),
        valid_from=valid_from,
        valid_to=valid_to,
        all_in_flag=True,
        notes_summary=None,
    )

    offers: list[RateOffer] = []
    charges: list[RateChargeLine] = []
    notes = extract_notes(workbook, template, card.id)

    header_row = template.header_detection.get("fixed_header_row", 1)
    multi_row_header = template.header_detection.get("multi_row_header", False)
    for sheet_name in matched_sheets:
        sheet = workbook[sheet_name]
        header_rows = [row for row in sheet.iter_rows(min_row=header_row, max_row=header_row + (1 if multi_row_header else 0), values_only=True)]
        headers = combine_headers(header_rows)
        field_indexes = {field: find_header(headers, label) for field, label in template.field_map.items()}
        charge_indexes = []
        for column in template.breakdown_columns:
            index = find_header(headers, column["source_label"])
            if index is not None:
                charge_indexes.append((column, index))

        for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + (2 if multi_row_header else 1) + 1, values_only=True), start=header_row + (2 if multi_row_header else 1) + 1):
            values = [normalize_text(value) for value in row]
            if should_skip_row(values, template.row_filters):
                continue

            equipment_raw = value_at(row, field_indexes.get("equipment_type"))
            equipment_type, _ = normalize_equipment(str(equipment_raw or ""), template.normalizers.get("equipment_type", {}))
            amount, routing_tail = parse_amount(value_at(row, field_indexes.get("base_amount")))
            if not equipment_type or amount is None:
                continue

            offer = RateOffer(
                rate_card_id=card.id,
                zone=clean_nullable(normalize_text(value_at(row, field_indexes.get("zone")))),
                pol=normalize_text(value_at(row, field_indexes.get("pol"))) or None,
                pod=normalize_text(value_at(row, field_indexes.get("pod"))) or None,
                final_destination=clean_nullable(normalize_text(value_at(row, field_indexes.get("final_destination")))),
                equipment_type=equipment_type,
                base_amount=amount,
                base_currency=normalize_text(value_at(row, field_indexes.get("base_currency"))) or card.currency_default,
                all_in_flag=True,
                routing_note=routing_tail,
                valid_from=parse_date_range(value_at(row, field_indexes.get("valid_from")))[0] or card.valid_from,
                valid_to=parse_date_range(value_at(row, field_indexes.get("valid_to")))[1] or card.valid_to,
                raw_sheet_name=sheet_name,
                raw_row_reference=f"{sheet_name}!R{row_number}",
                raw_row_json={headers[i] if i < len(headers) else f"col_{i+1}": values[i] for i in range(len(values))},
            )
            offers.append(offer)

            for charge_cfg, index in charge_indexes:
                charge_amount, _ = parse_amount(value_at(row, index))
                if charge_amount is None:
                    continue
                charges.append(
                    RateChargeLine(
                        rate_offer_id=offer.id,
                        charge_name=charge_cfg["charge_name"],
                        charge_type=charge_cfg.get("charge_type"),
                        amount=charge_amount,
                        currency=card.currency_default,
                        source_label=charge_cfg["source_label"],
                        raw_value=normalize_text(value_at(row, index)),
                    )
                )
    if notes:
        card.notes_summary = notes[0].note_text[:240]
    return card, offers, charges, notes


def combine_headers(rows: list[tuple[object, ...]]) -> list[str]:
    width = max((len(row) for row in rows), default=0)
    combined = []
    for index in range(width):
        parts = []
        for row in rows:
            value = normalize_text(row[index] if index < len(row) else "")
            if value and value not in parts:
                parts.append(value)
        combined.append(" ".join(parts))
    return combined


def find_header(headers: list[str], label: str) -> int | None:
    needle = normalize_text(label).upper()
    for index, header in enumerate(headers):
        normalized = normalize_text(header).upper()
        if needle and needle in normalized:
            return index
    return None


def value_at(row: tuple[object, ...], index: int | None) -> object | None:
    if index is None or index >= len(row):
        return None
    return row[index]


def should_skip_row(values: list[str], row_filters: dict) -> bool:
    if row_filters.get("skip_if_first_cell_empty") and not values[:1]:
        return True
    if row_filters.get("skip_if_first_cell_empty") and values and not values[0]:
        return True
    skip_tokens = [token.lower() for token in row_filters.get("skip_rows_containing", [])]
    text = " ".join(values).lower()
    return any(token in text for token in skip_tokens)


def extract_metadata(workbook) -> dict[str, str]:
    if "Cover page" not in workbook.sheetnames:
        return {}
    sheet = workbook["Cover page"]
    metadata = {}
    for row in sheet.iter_rows(values_only=True):
        values = [normalize_text(value) for value in row if normalize_text(value)]
        if len(values) >= 2:
            metadata[values[0]] = values[-1]
    return metadata


def extract_notes(workbook, template: ParserTemplate, rate_card_id: str) -> list[RateNote]:
    notes = []
    note_sheets = template.note_extraction.get("scan_sheets", [])
    keywords = [token.lower() for token in template.note_extraction.get("keywords", [])]
    scan_top_rows = template.note_extraction.get("scan_top_rows", 20)
    for sheet_name in workbook.sheetnames:
        upper = sheet_name.upper()
        if note_sheets and not any(token.upper() in upper for token in note_sheets):
            continue
        sheet = workbook[sheet_name]
        for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(scan_top_rows, sheet.max_row), values_only=True), start=1):
            text = " ".join(filter(None, [normalize_text(value) for value in row]))
            if text and (not keywords or any(keyword in text.lower() for keyword in keywords)):
                notes.append(
                    RateNote(
                        rate_card_id=rate_card_id,
                        note_type="commercial",
                        note_text=text,
                        source_reference=f"{sheet_name}!R{row_number}",
                    )
                )
    return notes


def clean_nullable(value: str) -> str | None:
    if not value:
        return None
    if value.strip().upper() in {"N/A", "NA", "N.A."}:
        return None
    return value
