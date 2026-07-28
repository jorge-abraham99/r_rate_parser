from __future__ import annotations

import csv
import re
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from rate_ingest.models import ParserTemplate, RateCard, RateChargeLine, RateImport, RateNote, RateOffer
from rate_ingest.normalize import normalize_text

PORT_CODE_LABELS = {
    "GBFXT": "Felixstowe",
    "GBSOU": "Southampton",
    "GBLIV": "Liverpool",
    "GBLGP": "Tilbury",
    "GBTIL": "Tilbury",
    "GBTEE": "Teesport",
    "GBSSH": "Sheerness",
    "GBGRG": "Greenock",
    "GBGRK": "Grangemouth",
    "GBIMM": "Immingham",
}


def parse_workbook(
    path: Path, template: ParserTemplate, rate_import: RateImport
) -> tuple[RateCard, list[RateOffer], list[RateChargeLine], list[RateNote]]:
    valid_from, valid_to = infer_quarter_validity(path.name)
    card = RateCard(
        rate_import_id=rate_import.id,
        provider_name=template.provider_name,
        carrier_name=template.defaults.get("carrier_name", template.provider_name),
        document_type=template.document_type,
        commodity=template.defaults.get("commodity"),
        currency_default=template.defaults.get("currency_default", "GBP"),
        valid_from=valid_from,
        valid_to=valid_to,
        all_in_flag=False,
        notes_summary=None,
    )

    offers: list[RateOffer] = []
    notes: list[RateNote] = []
    charges: list[RateChargeLine] = []

    if path.suffix.lower() == ".csv":
        rows = load_csv_rows(path)
        offers.extend(parse_sheet_rows("csv_haulage", rows, card, valid_from, valid_to, path.name))
    else:
        workbook = load_workbook(path, data_only=True, read_only=True)
        for sheet_name in workbook.sheetnames:
            rows = [[cell for cell in row] for row in workbook[sheet_name].iter_rows(values_only=True)]
            offers.extend(parse_sheet_rows(sheet_name, rows, card, valid_from, valid_to, path.name))

    if offers:
        notes.append(
            RateNote(
                rate_card_id=card.id,
                note_type="commercial",
                note_text=f"Haulage tariff import from {path.name}",
                source_reference=path.name,
            )
        )
        card.notes_summary = notes[0].note_text
    return card, offers, charges, notes


def parse_sheet_rows(
    sheet_name: str,
    rows: list[list[object]],
    card: RateCard,
    valid_from: date | None,
    valid_to: date | None,
    source_name: str,
) -> list[RateOffer]:
    normalized_sheet = normalize_text(sheet_name).upper()
    if "GATEWAY" in normalized_sheet or "TILBURY" in normalized_sheet:
        return parse_gateway_sheet(sheet_name, rows, card, valid_from, valid_to)
    return parse_matrix_sheet(sheet_name, rows, card, valid_from, valid_to, source_name)


def parse_matrix_sheet(
    sheet_name: str,
    rows: list[list[object]],
    card: RateCard,
    valid_from: date | None,
    valid_to: date | None,
    source_name: str,
) -> list[RateOffer]:
    if len(rows) < 3:
        return []
    header = rows[1]
    port_columns: list[tuple[int, str, str]] = []
    for index, value in enumerate(header[1:], start=1):
        code = normalize_port_code(value)
        if not code:
            continue
        port_columns.append((index, code, port_label(code)))

    offers: list[RateOffer] = []
    for row_index, row in enumerate(rows[2:], start=3):
        collection = clean_nullable(row[0] if row else None)
        if not collection:
            continue
        for column_index, port_code, port_name in port_columns:
            amount = to_float(row[column_index] if column_index < len(row) else None)
            if amount is None or amount <= 0:
                continue
            offers.append(
                RateOffer(
                    rate_card_id=card.id,
                    offer_reference=f"{sheet_name}:{port_code}",
                    commodity=card.commodity,
                    origin=collection,
                    place_of_receipt=collection,
                    pol=port_name,
                    pod=port_name,
                    final_destination=port_name,
                    equipment_type="40HC",
                    service_mode="Door -> CY",
                    transit_time_days=None,
                    base_amount=round(amount, 2),
                    base_currency=card.currency_default,
                    all_in_flag=False,
                    valid_from=valid_from,
                    valid_to=valid_to,
                    raw_sheet_name=sheet_name,
                    raw_row_reference=f"{sheet_name}!R{row_index}C{column_index + 1}",
                    raw_row_json={
                        "collection_place": collection,
                        "port_code": port_code,
                        "source_file_name": source_name,
                    },
                )
            )
    return offers


def parse_gateway_sheet(
    sheet_name: str,
    rows: list[list[object]],
    card: RateCard,
    valid_from: date | None,
    valid_to: date | None,
) -> list[RateOffer]:
    offers: list[RateOffer] = []
    port_name = "Tilbury"
    for row_index, row in enumerate(rows[2:], start=3):
        collection = clean_nullable(row[0] if row else None)
        amount = to_float(row[1] if len(row) > 1 else None)
        if not collection or amount is None or amount <= 0:
            continue
        offers.append(
            RateOffer(
                rate_card_id=card.id,
                offer_reference=f"{sheet_name}:GBLGP-GBTIL",
                commodity=card.commodity,
                origin=collection,
                place_of_receipt=collection,
                pol=port_name,
                pod=port_name,
                final_destination=port_name,
                equipment_type="40HC",
                service_mode="Door -> CY",
                transit_time_days=None,
                base_amount=round(amount, 2),
                base_currency=card.currency_default,
                all_in_flag=False,
                valid_from=valid_from,
                valid_to=valid_to,
                raw_sheet_name=sheet_name,
                raw_row_reference=f"{sheet_name}!R{row_index}",
                raw_row_json={"collection_place": collection, "port_code": "GBLGP/GBTIL"},
            )
        )
    return offers


def infer_quarter_validity(file_name: str) -> tuple[date | None, date | None]:
    text = normalize_text(file_name).upper()
    match = re.search(r"\bQ([1-4])\b.*?\b(20\d{2})\b", text)
    if not match:
        return None, None
    quarter = int(match.group(1))
    year = int(match.group(2))
    month_start = (quarter - 1) * 3 + 1
    month_end = month_start + 2
    start = date(year, month_start, 1)
    if month_end == 12:
        end = date(year, 12, 31)
    else:
        next_month = date(year, month_end + 1, 1)
        end = next_month.fromordinal(next_month.toordinal() - 1)
    return start, end


def load_csv_rows(path: Path) -> list[list[object]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return [list(row) for row in csv.reader(handle)]


def normalize_port_code(value: object) -> str:
    text = normalize_text(value).upper().replace(" ", "")
    return text if text.startswith("GB") else ""


def port_label(code: str) -> str:
    return PORT_CODE_LABELS.get(code, code)


def clean_nullable(value: object) -> str | None:
    text = normalize_text(value)
    return text or None


def to_float(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        text = normalize_text(value).replace(",", "")
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None
