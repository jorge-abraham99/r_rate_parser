from __future__ import annotations

from pathlib import Path
from typing import Any
import csv

from openpyxl import load_workbook

from rate_ingest.email_source import dataframe_preview, load_email_payload, provider_from_sender
from rate_ingest.models import InspectResult, SourceDocument


def inspect_source(source_document: SourceDocument) -> InspectResult:
    source_path = Path(source_document.source_path)
    source_type = source_document.source_type.lower()
    if source_type == "eml":
        payload = load_email_payload(source_path)
        sheet_summaries = []
        for index, table in enumerate(payload["tables"], start=1):
            sheet_summaries.append(
                {
                    "sheet_name": f"email_table_{index}",
                    "dimensions": f"{len(table.index)} rows x {len(table.columns)} columns",
                    "top_rows": dataframe_preview(table),
                }
            )
        provider_guess = provider_from_name(source_document.file_name) or provider_from_sender(payload["sender"])
        return InspectResult(
            source_document=source_document,
            workbook_type=source_type,
            provider_guess=provider_guess,
            parser_family_guess=guess_parser_family(sheet_summaries, source_type=source_type),
            sheet_summaries=sheet_summaries,
        )

    if source_type == "csv":
        rows: list[list[str]] = []
        with source_path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            for index, row in enumerate(reader):
                values = [normalize_cell(value) for value in row]
                if any(values):
                    rows.append(values[:12])
                if index >= 9:
                    break
        provider_guess = provider_from_name(source_document.file_name)
        return InspectResult(
            source_document=source_document,
            workbook_type=source_type,
            provider_guess=provider_guess,
            parser_family_guess=guess_parser_family([{"sheet_name": "csv_1", "top_rows": rows}], source_type=source_type),
            sheet_summaries=[{"sheet_name": "csv_1", "dimensions": "csv", "top_rows": rows}],
        )

    if source_type == "pdf":
        try:
            import pymupdf as fitz
        except ImportError as exc:
            raise ValueError("PDF inspection requires the pymupdf package.") from exc
        sheet_summaries = []
        document_text: list[str] = []
        with fitz.open(source_path) as document:
            for page_number, page in enumerate(document, start=1):
                page_text = page.get_text("text", sort=True)
                document_text.append(page_text)
                lines = [normalize_cell(line) for line in page_text.splitlines() if normalize_cell(line)]
                sheet_summaries.append(
                    {
                        "sheet_name": f"page_{page_number}",
                        "dimensions": f"{round(page.rect.width)} x {round(page.rect.height)} points",
                        "top_rows": [[line] for line in lines],
                    }
                )
        combined_text = " ".join(document_text)
        return InspectResult(
            source_document=source_document,
            workbook_type=source_type,
            provider_guess=provider_from_name(f"{source_document.file_name} {combined_text}"),
            parser_family_guess=guess_parser_family(sheet_summaries, source_type=source_type),
            sheet_summaries=sheet_summaries,
        )

    if source_type not in {"xlsx", "xlsm", "xls"}:
        return InspectResult(
            source_document=source_document,
            workbook_type=source_type,
            provider_guess=provider_from_name(source_document.file_name),
        )

    workbook = load_workbook(source_path, data_only=True, read_only=True)
    # Some legacy Maersk workbooks use a layout that openpyxl's streaming
    # reader reports as a 1x1 sheet. Fall back to the normal reader for those
    # files while keeping large AFLS workbooks streaming-friendly.
    if any(
        sheet.max_row <= 1 and sheet.max_column <= 1
        for sheet in workbook.worksheets
    ):
        workbook.close()
        workbook = load_workbook(source_path, data_only=True, read_only=False)
    sheet_summaries = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        top_rows = []
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 25), values_only=True):
            values = [normalize_cell(value) for value in row]
            if any(values):
                top_rows.append(values[:32])
        sheet_summaries.append(
            {
                "sheet_name": sheet_name,
                "dimensions": f"{sheet.max_row} rows x {sheet.max_column} columns",
                "top_rows": top_rows,
            }
        )

    parser_guess = guess_parser_family(sheet_summaries, source_type=source_type)
    return InspectResult(
        source_document=source_document,
        workbook_type=source_type,
        provider_guess=provider_from_name(source_document.file_name),
        parser_family_guess=parser_guess,
        sheet_summaries=sheet_summaries,
    )


def provider_from_name(file_name: str) -> str | None:
    upper = file_name.upper()
    if "HAPAG" in upper:
        return "HAPAG-LLOYD"
    if "CMA" in upper:
        return "CMA CGM"
    if "QT-MAEU" in upper or "MAEU" in upper:
        return "MAERSK"
    for provider in ("MSC", "COSCO", "MAERSK"):
        if provider in upper:
            return provider
    return None


def guess_parser_family(sheet_summaries: list[dict[str, Any]], source_type: str | None = None) -> str | None:
    sheet_names = " ".join(summary.get("sheet_name", "") for summary in sheet_summaries).upper()
    flattened = " ".join(
        " ".join(" ".join(row) for row in summary.get("top_rows", [])) for summary in sheet_summaries
    ).upper()
    if (
        source_type == "pdf"
        and "COSCO" in flattened
        and "FREIGHT RATE" in flattened
        and "EMERGENCY FUEL SURCHARGE" in flattened
        and "INLAND HAULAGE AT" in flattened
    ):
        return "cosco_pdf_quote"
    if source_type == "eml" and "POO" in flattened and "POL/POD" in flattened and "OFFER GIGO" in flattened:
        return "email_table"
    if "CITY NAME" in flattened and ("GBFXT" in flattened or "GBSOU" in flattened or "GBLGP" in flattened):
        return "haulage_matrix"
    has_msc_zoned_rate_pair = (
        "REUDAN-SPECIAL" in sheet_names
        and "REUDAN-TARRIFF" in sheet_names
    ) or (
        "HAULAGE ZONES SEP" in sheet_names
        and "REUDAN-PEUTE" in sheet_names
        and "REUDAN-PAPER" in sheet_names
    )
    if (
        "HAULAGE ZONES" in sheet_names
        and has_msc_zoned_rate_pair
        and "CITY" in flattened
        and "COUNTY" in flattened
        and "ALL IN RATE" in flattened
    ):
        return "msc_zoned_inline"
    if (
        "REUDAN-PEUTE" in sheet_names
        and "REUDAN-PAPER" in sheet_names
        and "CUSTOMER" in flattened
        and "POL" in flattened
        and "POD" in flattened
    ):
        return "tabular_lane"
    if (
        "SRV ID" in flattened
        and "CHARGE TYPE" in flattened
        and "CHARGE CODE" in flattened
        and "UNIT OF MEASURE" in flattened
        and "PORT OF LOADING" in flattened
        and "PORT OF DISCHARGE" in flattened
    ):
        return "hapag_india_rows"
    if (
        "GEO FROM STD LOCATION" in flattened
        and "PREFERRED POL" in flattened
        and "APPLICABLE ROUTING" in flattened
        and "ALL IN RATE" in flattened
    ):
        return "hapag_door_matrix"
    if "RECEIPT" in flattened and "DELIVERY" in flattened and "COMMODITY NAME" in flattened and "RATE BASIS" in flattened:
        return "site_to_site_rows"
    if "OFFER 1-1" in flattened or "SCHEDULED ROUTE" in flattened:
        return "offer_block"
    if "CUSTOMER" in flattened and "POL" in flattened and "POD" in flattened:
        return "tabular_lane"
    if "TERMS AND CONDITIONS - POL" in flattened or "VIA SOU" in flattened:
        return "matrix"
    return "unknown"


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("\n", " / ")
