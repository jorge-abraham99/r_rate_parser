from __future__ import annotations

import json
import sys
from pathlib import Path
from uuid import UUID

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from typer.testing import CliRunner

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from rate_ingest.api import app as api_app
from rate_ingest.auth import (
    AuthenticatedUser,
    OrganizationMembership,
    RequestContext,
    require_operator,
    require_organization_member,
)
from rate_ingest.cli import app
from rate_ingest.config import Settings
from rate_ingest.models import InspectResult, SourceDocument
from rate_ingest.template_matcher import load_templates, score_template


runner = CliRunner()
api_client = TestClient(api_app)


@pytest.fixture(autouse=True)
def authenticated_api_client():
    context = RequestContext(
        user=AuthenticatedUser(
            user_id=UUID("123e4567-e89b-12d3-a456-426614174000"),
            email="test@example.com",
            access_token="test-token",
            claims={},
        ),
        memberships=(
            OrganizationMembership(
                organization_id=UUID("123e4567-e89b-12d3-a456-426614174001"),
                organization_name="Reudan",
                organization_slug="reudan",
                role="admin",
            ),
        ),
    )
    api_app.dependency_overrides[require_organization_member] = lambda: context
    api_app.dependency_overrides[require_operator] = lambda: context
    yield
    api_app.dependency_overrides.clear()


def seed_templates(tmp_path: Path) -> None:
    templates_dir = tmp_path / "data" / "templates"
    templates_dir.mkdir(parents=True)
    for template_path in Path("data/templates").glob("*.yaml"):
        templates_dir.joinpath(template_path.name).write_text(
            template_path.read_text(encoding="utf-8"),
            encoding="utf-8",
        )


def build_msc_peute_paper_fixture(tmp_path: Path) -> Path:
    source = tmp_path / "Reudan - Export Far East - Waste Paper Rate Sheet - September 2026.xlsx"
    workbook = load_workbook("rate_sheet_files/MSC - FAR EAST  AUGUST.xlsx")
    workbook["REUDAN-SPECIAL"].title = "REUDAN-PEUTE"
    workbook["REUDAN-TARRIFF"].title = "REUDAN-PAPER"
    workbook["Haulage Zones"].title = "Haulage Zones SEP"

    for sheet_name, documentation in (
        ("REUDAN-PEUTE", "USD 40 per B/L"),
        ("REUDAN-PAPER", "GBP 30 per B/L"),
    ):
        sheet = workbook[sheet_name]
        for row_number in range(4, 256):
            sheet.cell(row=row_number, column=7).value = documentation
            sheet.cell(row=row_number, column=9).value = "01.09.26"
            sheet.cell(row=row_number, column=10).value = "30.09.26"

    cover = workbook["Cover page"]
    for row in cover.iter_rows():
        if any(cell.value == "Validity From/To" for cell in row):
            next(cell for cell in reversed(row) if cell.value is not None).value = "01.09.26 - 30.09.26"
            break

    haulage = workbook["Haulage Zones SEP"]
    duplicate = [haulage.cell(row=3, column=column).value for column in range(1, 6)]
    haulage.append(duplicate)
    workbook.save(source)
    return source


def test_settings_seed_missing_bundled_templates_without_overwriting_existing(tmp_path: Path, monkeypatch):
    monkeypatch.delenv("RATE_INGEST_ROOT", raising=False)
    template_dir = tmp_path / "data" / "templates"
    template_dir.mkdir(parents=True)
    existing = template_dir / "msc_far_east_v1.yaml"
    existing.write_text("operator-managed template", encoding="utf-8")

    settings = Settings.load(cwd=tmp_path)
    settings.ensure()

    assert existing.read_text(encoding="utf-8") == "operator-managed template"
    assert settings.templates_dir.joinpath("msc_zoned_inline_v1.yaml").exists()
    assert {
        path.name for path in settings.templates_dir.glob("*.yaml")
    } == {
        path.name for path in Path("data/templates").glob("*.yaml")
    }


def test_cosco_pdf_template_loads_from_bundle_and_strongly_matches_tuticorin(tmp_path: Path):
    settings = Settings.load(cwd=tmp_path)
    settings.templates_dir.mkdir(parents=True)
    template = next(
        item for item in load_templates(settings)
        if item.template_id == "cosco_pdf_quote_v1"
    )
    source = SourceDocument(
        source_type="pdf",
        file_name="Tuticorin.pdf",
        source_path=str(tmp_path / "Tuticorin.pdf"),
        checksum="test",
    )
    inspected = InspectResult(
        source_document=source,
        workbook_type="pdf",
        provider_guess=None,
        parser_family_guess="unknown",
        sheet_summaries=[],
    )

    assert score_template(template, inspected) >= 0.55


def test_inspect_and_import_and_approve_flow(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    raw_dir = tmp_path / "incoming"
    raw_dir.mkdir()
    source = raw_dir / "MSC - FAR EAST RATES JAN.xlsx"
    source.write_bytes(Path("rate_sheet_files/MSC - FAR EAST RATES JAN.xlsx").read_bytes())
    seed_templates(tmp_path)

    result = runner.invoke(app, ["inspect", str(source)])
    assert result.exit_code == 0
    assert "Likely parser family" in result.stdout

    result = runner.invoke(app, ["import", str(source)])
    assert result.exit_code == 0
    assert "Import created:" in result.stdout
    import_id = next(line.split(": ", 1)[1] for line in result.stdout.splitlines() if line.startswith("Import created:"))
    run_dir = tmp_path / "data" / "runs" / import_id
    assert run_dir.joinpath("canonical_rates.json").exists()
    canonical_rates = json.loads(run_dir.joinpath("canonical_rates.json").read_text(encoding="utf-8"))
    assert canonical_rates
    assert set(canonical_rates[0].keys()) == {
        "rate_type",
        "from_raw",
        "to_raw",
        "amount",
        "currency",
        "unit",
        "valid_from",
        "valid_to",
    }

    result = runner.invoke(app, ["approve", import_id, "--approved-by", "jorge"])
    assert result.exit_code == 0
    approved_rates = (tmp_path / "data" / "warehouse" / "approved_rates.csv").read_text(encoding="utf-8")
    assert "rate_type,from_raw,to_raw,amount,currency,unit,valid_from,valid_to" in approved_rates

    result = runner.invoke(app, ["search", "--pod", "HO CHI MINH"])
    assert result.exit_code == 0
    assert "MSC" in result.stdout


def test_cosco_matrix_import_creates_canonical_rates(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    raw_dir = tmp_path / "incoming"
    raw_dir.mkdir()
    source = raw_dir / "COSCO FAR-EAST RATES.xlsx"
    source.write_bytes(Path("rate_sheet_files/COSCO FAR-EAST RATES.xlsx").read_bytes())
    seed_templates(tmp_path)

    result = runner.invoke(app, ["import", str(source)])
    assert result.exit_code == 0
    assert "Template used: cosco_matrix_v1" in result.stdout
    import_id = next(line.split(": ", 1)[1] for line in result.stdout.splitlines() if line.startswith("Import created:"))
    run_dir = tmp_path / "data" / "runs" / import_id
    canonical_rates = json.loads(run_dir.joinpath("canonical_rates.json").read_text(encoding="utf-8"))
    assert canonical_rates
    first = canonical_rates[0]
    assert first["rate_type"] == "ocean"
    assert first["unit"] == "per_container"
    assert first["from_raw"]
    assert first["to_raw"]


def test_cosco_csv_import_extracts_efs_as_charge_lines(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    raw_dir = tmp_path / "incoming"
    raw_dir.mkdir()
    source = raw_dir / "COSCO September rates.csv"
    source.write_bytes(Path("rate_sheet_files/cosco_sep.csv").read_bytes())
    seed_templates(tmp_path)

    result = runner.invoke(app, ["import", str(source)])
    assert result.exit_code == 0
    assert "Template used: cosco_csv_quote_v1" in result.stdout
    import_id = next(
        line.split(": ", 1)[1]
        for line in result.stdout.splitlines()
        if line.startswith("Import created:")
    )
    run_dir = tmp_path / "data" / "runs" / import_id
    offers = detail_rows(run_dir / "parsed_rate_offers.csv")
    charges = detail_rows(run_dir / "parsed_rate_charge_lines.csv")

    assert len(offers) == 15
    assert len(charges) == 15
    assert {offer["pol"] for offer in offers} == {"Felixstowe", "Southampton"}
    assert {charge["amount"] for charge in charges} == {"100.0", "150.0"}
    assert {charge["charge_name"] for charge in charges} == {"Emergency Fuel Surcharge"}
    assert all(charge["currency"] == "USD" for charge in charges)
    assert any(offer["base_amount"] == "50.0" for offer in offers)


def test_cosco_haulage_import_uses_streaming_header_and_canonical_ports(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    raw_dir = tmp_path / "incoming"
    raw_dir.mkdir()
    source = raw_dir / "COSCO Haulage - sep.xlsx"
    source.write_bytes(Path("rate_sheet_files/Cosco Haulage - sep.xlsx").read_bytes())
    seed_templates(tmp_path)

    result = runner.invoke(app, ["import", str(source)])
    assert result.exit_code == 0
    assert "Template used: cosco_haulage_v1" in result.stdout
    import_id = next(
        line.split(": ", 1)[1]
        for line in result.stdout.splitlines()
        if line.startswith("Import created:")
    )
    run_dir = tmp_path / "data" / "runs" / import_id
    offers = detail_rows(run_dir / "parsed_rate_offers.csv")

    assert len(offers) == 3630
    assert {offer["pol"] for offer in offers} == {"Felixstowe", "Southampton"}
    assert {offer["equipment_type"] for offer in offers} == {"40HC"}
    assert {offer["base_currency"] for offer in offers} == {"USD"}
    assert offers[0]["service_mode"] == "Door -> CY"


def test_cosco_ocean_and_haulage_publish_as_separate_current_sources(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    raw_dir = tmp_path / "incoming"
    raw_dir.mkdir()
    ocean_source = raw_dir / "COSCO September rates.csv"
    haulage_source = raw_dir / "COSCO Haulage - sep.xlsx"
    ocean_source.write_bytes(Path("rate_sheet_files/cosco_sep.csv").read_bytes())
    haulage_source.write_bytes(Path("rate_sheet_files/Cosco Haulage - sep.xlsx").read_bytes())
    seed_templates(tmp_path)

    ocean_result = runner.invoke(app, ["import", str(ocean_source)])
    haulage_result = runner.invoke(app, ["import", str(haulage_source)])
    assert ocean_result.exit_code == 0
    assert haulage_result.exit_code == 0
    ocean_id = next(
        line.split(": ", 1)[1]
        for line in ocean_result.stdout.splitlines()
        if line.startswith("Import created:")
    )
    haulage_id = next(
        line.split(": ", 1)[1]
        for line in haulage_result.stdout.splitlines()
        if line.startswith("Import created:")
    )

    for import_id, key, label, tag in (
        (ocean_id, "cosco-sea", "COSCO · Quay-to-quay", "SEA"),
        (haulage_id, "cosco-haulage", "COSCO · Export haulage", "HAUL"),
    ):
        response = api_client.post(
            f"/api/imports/{import_id}/approve",
            json={
                "approved_by": "jorge",
                "carrier_name": "COSCO",
                "carrier_key": key,
                "carrier_label": label,
                "contract_tag": tag,
            },
        )
        assert response.status_code == 200

    statuses = {
        item["import_id"]: item["status"]
        for item in api_client.get("/api/imports").json()
        if item["import_id"] in {ocean_id, haulage_id}
    }
    assert statuses == {ocean_id: "approved", haulage_id: "approved"}
    desk = api_client.get("/api/rate-desk", params={"limit": 50}).json()
    assert len(desk["rates"]) == 15
    fifty_dollar_rate = next(rate for rate in desk["rates"] if rate["base_amount"] == 50.0)
    assert fifty_dollar_rate["all_in_usd"] == 150.0
    assert desk["haulage_tariffs"]["abercarn caerphilly wales united kingdom"]["felixstowe"] == 364.55
    assert desk["haulage_tariffs_by_source"]["cosco-haulage"]["abercarn caerphilly wales united kingdom"]["felixstowe"] == 364.55


def test_cosco_pdf_quote_prices_only_freight_efs_and_haulage(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    raw_dir = tmp_path / "incoming"
    raw_dir.mkdir()
    source = raw_dir / "Tuticorin.pdf"
    source.write_bytes(Path("rate_sheet_files/Tuticorin.pdf").read_bytes())
    seed_templates(tmp_path)

    result = runner.invoke(app, ["import", str(source)])
    assert result.exit_code == 0
    assert "Template used: cosco_pdf_quote_v1" in result.stdout
    import_id = next(line.split(": ", 1)[1] for line in result.stdout.splitlines() if line.startswith("Import created:"))
    run_dir = tmp_path / "data" / "runs" / import_id

    offers = detail_rows(run_dir / "parsed_rate_offers.csv")
    charges = detail_rows(run_dir / "parsed_rate_charge_lines.csv")
    assert len(offers) == 76
    assert len(charges) == 152
    assert {offer["equipment_type"] for offer in offers} == {"40GP", "40HC"}
    assert {offer["service_mode"] for offer in offers} == {"SD / CY"}
    assert {offer["valid_from"] for offer in offers} == {"2026-05-01"}
    assert {offer["valid_to"] for offer in offers} == {"2026-05-31"}
    assert {charge["charge_name"] for charge in charges} == {
        "Emergency Fuel Surcharge",
        "Inland Haulage at Load",
    }
    assert not any(
        token in charge["charge_name"].lower()
        for charge in charges
        for token in ("documentation", "terminal handling")
    )

    birmingham = next(
        offer
        for offer in offers
        if offer["place_of_receipt"] == "Birmingham" and offer["equipment_type"] == "40HC"
    )
    assert birmingham["pol"] == "Felixstowe"
    assert birmingham["pod"] == "Tuticorin"
    assert float(birmingham["base_amount"]) == 360.0
    birmingham_charges = {
        charge["charge_name"]: float(charge["amount"])
        for charge in charges
        if charge["rate_offer_id"] == birmingham["id"]
    }
    assert birmingham_charges == {
        "Emergency Fuel Surcharge": 150.0,
        "Inland Haulage at Load": 264.0,
    }

    approve_response = api_client.post(
        f"/api/imports/{import_id}/approve",
        json={
            "approved_by": "jorge",
            "carrier_name": "COSCO",
            "carrier_key": "cosco-door",
            "carrier_label": "COSCO · India/Far East Door-to-quay",
        },
    )
    assert approve_response.status_code == 200
    search = api_client.get(
        "/api/search",
        params={
            "collection": "Birmingham",
            "pod": "Tuticorin",
            "equipment_type": "40HC",
            "limit": 20,
        },
    ).json()
    assert len(search) == 1
    assert search[0]["base_amount"] == 360.0
    assert search[0]["charge_total"] == 414.0
    assert search[0]["all_in_amount"] == 774.0
    assert search[0]["carrier_label"] == "COSCO · India/Far East Door-to-quay"

    desk = api_client.get("/api/rate-desk", params={"limit": 1000}).json()
    assert desk["haulage_tariffs"] == {}
    assert desk["filters"]["door_pickups"] == []
    assert "Birmingham, GB" in desk["filters"]["collection_places"]


def test_maersk_offer_block_import_creates_charge_lines(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    raw_dir = tmp_path / "incoming"
    raw_dir.mkdir()
    source = raw_dir / "MAERSK Q-1, INDIA AND FAR-EAST.xlsx"
    source.write_bytes(Path("rate_sheet_files/MAERSK Q-1, INDIA AND FAR-EAST.xlsx").read_bytes())
    seed_templates(tmp_path)

    result = runner.invoke(app, ["import", str(source)])
    assert result.exit_code == 0
    assert "Template used: maersk_offer_block_v1" in result.stdout
    import_id = next(line.split(": ", 1)[1] for line in result.stdout.splitlines() if line.startswith("Import created:"))
    run_dir = tmp_path / "data" / "runs" / import_id
    canonical_rates = json.loads(run_dir.joinpath("canonical_rates.json").read_text(encoding="utf-8"))
    assert canonical_rates
    assert canonical_rates[0]["valid_from"] == "2025-12-01"
    assert canonical_rates[0]["valid_to"] == "2025-12-31"
    parsed_charges = run_dir.joinpath("parsed_rate_charge_lines.csv").read_text(encoding="utf-8")
    assert "Basic Ocean Freight" in parsed_charges


def test_maersk_qtmaeu_offer_block_autodetects_and_carries_material(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    raw_dir = tmp_path / "incoming"
    raw_dir.mkdir()
    source = raw_dir / "QT-MAEU-50875469-0.xlsx"
    source.write_bytes(Path("rate_sheet_files/Maersk Rates - Apr to June /QT-MAEU-50875469-0.xlsx").read_bytes())
    seed_templates(tmp_path)

    result = runner.invoke(app, ["import", str(source)])
    assert result.exit_code == 0
    assert "Template used: maersk_offer_block_v1" in result.stdout
    import_id = next(line.split(": ", 1)[1] for line in result.stdout.splitlines() if line.startswith("Import created:"))
    run_dir = tmp_path / "data" / "runs" / import_id
    canonical_rates = json.loads(run_dir.joinpath("canonical_rates.json").read_text(encoding="utf-8"))
    assert canonical_rates
    assert canonical_rates[0]["valid_from"] == "2026-04-01"
    assert canonical_rates[0]["valid_to"] == "2026-04-30"
    offers = [row for row in detail_rows(run_dir / "parsed_rate_offers.csv")]
    assert offers[0]["commodity"] == "WASTEPAPER"
    assert offers[0]["pol"] == "Antwerp"


def test_maersk_rate_desk_exposes_charge_analysis(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    seed_templates(tmp_path)
    source_bytes = Path("rate_sheet_files/MAERSK Q-1, INDIA AND FAR-EAST.xlsx").read_bytes()

    response = api_client.post(
        "/api/imports",
        data={"uploaded_by": "jorge"},
        files={"file": ("MAERSK Q-1, INDIA AND FAR-EAST.xlsx", source_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    import_id = response.json()["import_id"]

    detail_response = api_client.get(f"/api/imports/{import_id}")
    assert detail_response.status_code == 200
    detail = detail_response.json()
    assert detail["charge_bucket_summary"]["matched_charge_count"] > 0
    assert detail["charge_bucket_summary"]["unmatched_charge_count"] == 0
    assert [group["key"] for group in detail["charge_bucket_summary"]["groups"]] == ["origin", "freight", "destination"]

    approve_response = api_client.post(
        f"/api/imports/{import_id}/approve",
        json={
            "approved_by": "jorge",
            "carrier_name": "Maersk",
            "carrier_key": "maersk-demo",
            "carrier_label": "Maersk Demo",
            "contract_tag": "SPOT",
        },
    )
    assert approve_response.status_code == 200

    desk_response = api_client.get("/api/rate-desk")
    assert desk_response.status_code == 200
    desk = desk_response.json()
    maersk_rate = next(
        rate
        for rate in desk["rates"]
        if rate["source_file_name"] == "MAERSK Q-1, INDIA AND FAR-EAST.xlsx"
        and rate["offer_reference"] == "Offer 2-1"
    )
    analysis = maersk_rate["charge_analysis"]
    assert maersk_rate["transit_time_days"] == 51
    assert maersk_rate["valid_from"] == "2025-12-01"
    assert maersk_rate["valid_to"] == "2025-12-31"
    assert analysis["matched_charge_count"] > 0
    assert analysis["unmatched_charge_count"] == 0
    assert analysis["total_usd"] > 0
    assert [group["key"] for group in analysis["groups"]] == ["origin", "freight", "destination"]
    assert analysis["groups"][0]["subtotal_usd"] >= 0
    assert analysis["groups"][1]["subtotal_usd"] > 0
    assert analysis["groups"][2]["subtotal_usd"] >= 0
    assert maersk_rate["all_in_usd"] == analysis["total_usd"]


def test_rate_desk_summary_paginates_and_loads_offer_detail_on_demand(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    seed_templates(tmp_path)
    source_bytes = Path("rate_sheet_files/MSC - FAR EAST RATES JAN.xlsx").read_bytes()

    response = api_client.post(
        "/api/imports",
        data={"uploaded_by": "jorge"},
        files={
            "file": (
                "MSC - FAR EAST RATES JAN.xlsx",
                source_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200, response.text
    import_id = response.json()["import_id"]
    approve_response = api_client.post(
        f"/api/imports/{import_id}/approve",
        json={
            "approved_by": "jorge",
            "carrier_name": "MSC",
            "carrier_key": "msc-inline",
            "carrier_label": "MSC · Door-to-quay",
        },
    )
    assert approve_response.status_code == 200, approve_response.text

    metadata_response = api_client.get("/api/rate-desk/meta")
    assert metadata_response.status_code == 200
    metadata = metadata_response.json()
    assert "rates" not in metadata
    assert metadata["filters"]["destinations"]

    page_response = api_client.get(
        "/api/rate-desk/search",
        params={"limit": 1, "offset": 0},
    )
    assert page_response.status_code == 200
    page = page_response.json()
    assert len(page["rates"]) == 1
    assert page["pagination"]["total"] > 1
    summary = page["rates"][0]
    assert "charges" not in summary
    assert "charge_analysis" not in summary
    assert "notes" not in summary
    assert "zone" in summary
    assert summary["all_in_usd"] is not None

    carrier_page_response = api_client.get(
        "/api/rate-desk/search",
        params={"carrier_name": "MSC", "limit": 50, "offset": 0},
    )
    assert carrier_page_response.status_code == 200
    carrier_page = carrier_page_response.json()
    assert carrier_page["rates"]
    assert {rate["carrier_name"] for rate in carrier_page["rates"]} == {"MSC"}

    detail_response = api_client.get(
        f"/api/rate-desk/offers/{summary['offer_id']}"
    )
    assert detail_response.status_code == 200
    detail = detail_response.json()
    assert detail["offer_id"] == summary["offer_id"]
    assert detail["charges"]
    assert detail["charge_analysis"]["groups"]


def test_maersk_afls_site_to_site_import_creates_offers_and_charge_lines(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    raw_dir = tmp_path / "incoming"
    raw_dir.mkdir()
    source = raw_dir / "REUDAN_E1E_E3E_WAP_Q2 2026.xlsx"
    source.write_bytes(Path("rate_sheet_files/REUDAN_E1E_E3E_WAP_Q2 2026.xlsx").read_bytes())
    seed_templates(tmp_path)

    result = runner.invoke(app, ["import", str(source)])
    assert result.exit_code == 0
    assert "Template used: maersk_afls_site_to_site_v1" in result.stdout
    import_id = next(line.split(": ", 1)[1] for line in result.stdout.splitlines() if line.startswith("Import created:"))
    run_dir = tmp_path / "data" / "runs" / import_id
    canonical_rates = json.loads(run_dir.joinpath("canonical_rates.json").read_text(encoding="utf-8"))
    assert len(canonical_rates) > 1000
    first = canonical_rates[0]
    assert first["from_raw"] == "Alcester, GB"
    assert first["to_raw"] == "Bangkok, TH"
    assert first["amount"] == 450.0
    assert first["currency"] == "USD"
    parsed_charges = run_dir.joinpath("parsed_rate_charge_lines.csv").read_text(encoding="utf-8")
    assert "Documentation fee - Destination" in parsed_charges
    assert "Export Service" in parsed_charges
    offers = [row for row in detail_rows(run_dir / "parsed_rate_offers.csv")]
    assert offers[0]["commodity"] == "WASTEPAPER"


def test_maersk_september_india_total_uses_basic_freight_ebs_and_dti(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    seed_templates(tmp_path)
    source_bytes = Path("rate_sheet_files/september rates/MAERSK India.xlsx").read_bytes()

    response = api_client.post(
        "/api/imports",
        data={"uploaded_by": "jorge"},
        files={
            "file": (
                "MAERSK India.xlsx",
                source_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200, response.text
    import_id = response.json()["import_id"]
    detail = api_client.get(f"/api/imports/{import_id}").json()
    assert detail["rate_import"]["parser_family"] == "site_to_site_rows"
    assert detail["summary"]["rate_offers"] > 100
    assert all(
        offer["final_destination"].endswith(", IN")
        for offer in detail["offers_preview"]
    )

    first_offer = next(
        offer
        for offer in detail["offers_preview"]
        if offer["place_of_receipt"] == "Alcester, GB"
        and offer["final_destination"] == "Ennore Chennai, IN"
        and offer["equipment_type"] == "40HDRY"
    )
    assert {"BFS", "BAS", "EBS", "DTI"}.issubset(
        set(first_offer["raw_row_json"]["total_charge_codes"])
    )

    approve_response = api_client.post(
        f"/api/imports/{import_id}/approve",
        json={
            "approved_by": "jorge",
            "carrier_name": "Maersk",
            "carrier_key": "maersk-india",
            "carrier_label": "Maersk · India rates",
            "contract_tag": "INDIA",
        },
    )
    assert approve_response.status_code == 200, approve_response.text

    desk = api_client.get("/api/rate-desk", params={"limit": 5000}).json()
    rate = next(
        item
        for item in desk["rates"]
        if item["source_file_name"] == "MAERSK India.xlsx"
        and item["place_of_receipt"] == "Alcester, GB"
        and item["final_destination"] == "Ennore Chennai, IN"
        and item["equipment_type"] == "40HDRY"
    )
    assert rate["all_in_usd"] == 910.0
    assert rate["all_in_usd"] == rate["charge_analysis"]["total_usd"]
    assert {charge["source_label"] for charge in rate["charges"]} >= {
        "BAS",
        "CFD",
        "EBS",
        "DTI",
    }
    analysis_lines = [
        line
        for group in rate["charge_analysis"]["groups"]
        for line in group["lines"]
    ]
    assert any(line["counts_toward_total"] is False for line in analysis_lines)


def test_maersk_afls_rate_desk_preserves_service_mode(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    seed_templates(tmp_path)
    source_bytes = Path("rate_sheet_files/REUDAN_E1E_E3E_WAP_Q2 2026.xlsx").read_bytes()

    response = api_client.post(
        "/api/imports",
        data={"uploaded_by": "jorge"},
        files={"file": ("REUDAN_E1E_E3E_WAP_Q2 2026.xlsx", source_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    import_id = response.json()["import_id"]

    approve_response = api_client.post(
        f"/api/imports/{import_id}/approve",
        json={
            "approved_by": "jorge",
            "carrier_name": "Maersk",
            "carrier_key": "maersk-contract",
            "carrier_label": "Maersk · Contract",
            "contract_tag": "CONTRACT",
        },
    )
    assert approve_response.status_code == 200

    desk = api_client.get("/api/rate-desk").json()
    row = next(
        item
        for item in desk["rates"]
        if item.get("place_of_receipt") == "Alcester, GB" and item.get("final_destination") == "Bangkok, TH"
    )
    assert row["service_mode"] == "SD / CY"


def test_haulage_matrix_import_autodetects_and_exposes_tariffs(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    raw_dir = tmp_path / "incoming"
    raw_dir.mkdir()
    source = raw_dir / "Export Waste Haulage Bristol + ALL other UK POLS INC GBLGP - GBTIL - Q2 2026 VALIDITY.xlsx"
    source.write_bytes(
        Path("rate_sheet_files/Export Waste Haulage Bristol + ALL other UK POLS INC GBLGP - GBTIL - Q2 2026 VALIDITY.xlsx").read_bytes()
    )
    seed_templates(tmp_path)

    result = runner.invoke(app, ["import", str(source)])
    assert result.exit_code == 0
    assert "Template used: uk_haulage_matrix_v1" in result.stdout
    import_id = next(line.split(": ", 1)[1] for line in result.stdout.splitlines() if line.startswith("Import created:"))
    run_dir = tmp_path / "data" / "runs" / import_id
    canonical_rates = json.loads(run_dir.joinpath("canonical_rates.json").read_text(encoding="utf-8"))
    assert canonical_rates
    assert canonical_rates[0]["rate_type"] == "inland_export"
    assert canonical_rates[0]["currency"] == "USD"
    assert canonical_rates[0]["valid_from"] == "2026-04-01"
    assert canonical_rates[0]["valid_to"] == "2026-06-30"

    approve_response = api_client.post(
        f"/api/imports/{import_id}/approve",
        json={
            "approved_by": "jorge",
            "carrier_name": "UK Inland Haulage",
            "carrier_key": "haulage-q2",
            "carrier_label": "UK Inland Haulage",
            "contract_tag": "HAUL",
        },
    )
    assert approve_response.status_code == 200

    desk = api_client.get("/api/rate-desk").json()
    assert desk["filters"]["door_pickups"]
    assert desk["haulage_currency"] == "USD"
    assert any(item["name"] == "ABBOTS BROMLEY" for item in desk["filters"]["door_pickups"])
    assert desk["haulage_tariffs"]["abbots bromley"]["felixstowe"] == 140.14
    assert desk["haulage_tariffs"]["abbots bromley"]["southampton"] == 105.21


def test_cma_email_import_creates_canonical_rates(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    raw_dir = tmp_path / "incoming"
    raw_dir.mkdir()
    source = raw_dir / "RE_ Far East Wastepaper for April - Reudan.eml"
    source.write_bytes(Path("RE_ Far East Wastepaper for April - Reudan.eml").read_bytes())
    seed_templates(tmp_path)

    result = runner.invoke(app, ["import", str(source)])
    assert result.exit_code == 0
    assert "Template used: cma_email_table_v1" in result.stdout
    import_id = next(line.split(": ", 1)[1] for line in result.stdout.splitlines() if line.startswith("Import created:"))
    run_dir = tmp_path / "data" / "runs" / import_id
    canonical_rates = json.loads(run_dir.joinpath("canonical_rates.json").read_text(encoding="utf-8"))
    assert canonical_rates
    first = canonical_rates[0]
    assert first["rate_type"] == "ocean"
    assert first["from_raw"] == "ACCRINGTON"
    assert first["currency"] == "USD"
    assert "MYPKG" in first["to_raw"]


def test_msc_zoned_inline_import_joins_birmingham_to_both_pols_and_tiers(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    raw_dir = tmp_path / "incoming"
    raw_dir.mkdir()
    source = raw_dir / "MSC - FAR EAST AUGUST.xlsx"
    source.write_bytes(Path("rate_sheet_files/MSC - FAR EAST  AUGUST.xlsx").read_bytes())
    seed_templates(tmp_path)

    result = runner.invoke(app, ["import", str(source)])
    assert result.exit_code == 0
    assert "Template used: msc_zoned_inline_v1" in result.stdout
    import_id = next(line.split(": ", 1)[1] for line in result.stdout.splitlines() if line.startswith("Import created:"))
    run_dir = tmp_path / "data" / "runs" / import_id

    matches = {}
    for offer in detail_rows(run_dir / "parsed_rate_offers.csv"):
        if offer["place_of_receipt"] != "Birmingham" or offer["pod"] != "SURABAYA":
            continue
        matches[(offer["offer_reference"], offer["pol"])] = offer

    assert {
        key: float(offer["base_amount"])
        for key, offer in matches.items()
    } == {
        ("SPECIAL", "FELIXSTOWE"): 650.0,
        ("SPECIAL", "LONDON GATEWAY"): 450.0,
        ("TARIFF", "FELIXSTOWE"): 665.0,
        ("TARIFF", "LONDON GATEWAY"): 465.0,
    }
    assert matches[("SPECIAL", "FELIXSTOWE")]["zone"] == "ZONE 3"
    assert matches[("SPECIAL", "LONDON GATEWAY")]["zone"] == "ZONE 2"
    assert all(offer["service_mode"] == "SD / CY" for offer in matches.values())

    tier_tables = json.loads(run_dir.joinpath("tier_rate_tables.json").read_text(encoding="utf-8"))
    assert len(tier_tables["SPECIAL"]) == 252
    assert len(tier_tables["TARIFF"]) == 252
    special_surabaya = next(
        row
        for row in tier_tables["SPECIAL"]
        if row["pol"] == "FELIXSTOWE" and row["zone"] == "ZONE 3" and "SURABAYA" in row["pod"]
    )
    assert special_surabaya["amount"] == 650.0
    assert special_surabaya["documentation"] == "GBP 30 per B/L"

    validation = json.loads(run_dir.joinpath("validation_report.json").read_text(encoding="utf-8"))
    assert validation["summary"]["errors"] == 0

    detail_response = api_client.get(f"/api/imports/{import_id}")
    assert detail_response.status_code == 200
    detail = detail_response.json()
    assert detail["card"]["valid_from"] == "2026-08-01"
    assert detail["card"]["valid_to"] == "2026-08-31"
    assert len(detail["tier_rate_tables"]["SPECIAL"]) == 252
    assert len(detail["tier_rate_tables"]["TARIFF"]) == 252

    approve_response = api_client.post(
        f"/api/imports/{import_id}/approve",
        json={
            "approved_by": "jorge",
            "carrier_name": "MSC",
            "carrier_key": "msc-inline",
            "carrier_label": "MSC · Door-to-quay",
        },
    )
    assert approve_response.status_code == 200
    search_response = api_client.get("/api/search", params={"pod": "SURABAYA", "limit": 5000})
    assert search_response.status_code == 200
    birmingham = {
        (rate["offer_reference"], rate["pol"]): rate
        for rate in search_response.json()
        if rate["place_of_receipt"] == "Birmingham"
    }
    assert birmingham[("SPECIAL", "FELIXSTOWE")]["all_in_amount"] == 680.0
    assert birmingham[("SPECIAL", "LONDON GATEWAY")]["all_in_amount"] == 480.0
    assert birmingham[("TARIFF", "FELIXSTOWE")]["all_in_amount"] == 695.0
    assert birmingham[("TARIFF", "LONDON GATEWAY")]["all_in_amount"] == 495.0

    desk_response = api_client.get("/api/rate-desk", params={"limit": 5000})
    assert desk_response.status_code == 200
    desk = desk_response.json()
    assert {"FELIXSTOWE", "LONDON GATEWAY"}.issubset(desk["filters"]["origins"])
    assert {"Surabaya, ID", "Semarang, ID"}.issubset(
        desk["filters"]["destinations"]
    )
    assert "Birmingham, GB" in desk["filters"]["collection_places"]
    assert desk["haulage_tariffs"] == {}
    assert desk["filters"]["door_pickups"] == []
    assert any(
        rate["carrier_label"] == "MSC · Door-to-quay" and rate["service_mode"] == "SD / CY"
        for rate in desk["rates"]
    )

    birmingham_response = api_client.get(
        "/api/search",
        params={"collection": "Birmingham", "pod": "SURABAYA", "limit": 5000},
    )
    assert birmingham_response.status_code == 200
    assert birmingham_response.json()
    assert all(rate["place_of_receipt"] == "Birmingham" for rate in birmingham_response.json())


def test_msc_peute_paper_import_applies_confirmed_documentation_rule(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    source = build_msc_peute_paper_fixture(tmp_path)
    seed_templates(tmp_path)

    result = runner.invoke(app, ["import", str(source)])
    assert result.exit_code == 0, result.stdout
    assert "Parser family: msc_zoned_inline" in result.stdout
    assert "Template used: msc_peute_paper_v1" in result.stdout
    import_id = next(
        line.split(": ", 1)[1]
        for line in result.stdout.splitlines()
        if line.startswith("Import created:")
    )
    run_dir = tmp_path / "data" / "runs" / import_id

    offers = detail_rows(run_dir / "parsed_rate_offers.csv")
    charges = detail_rows(run_dir / "parsed_rate_charge_lines.csv")
    tier_tables = json.loads(run_dir.joinpath("tier_rate_tables.json").read_text(encoding="utf-8"))
    validation = json.loads(run_dir.joinpath("validation_report.json").read_text(encoding="utf-8"))

    assert len(offers) > 30000
    assert len(charges) == len(offers)
    assert {offer["offer_reference"] for offer in offers} == {"PEUTE", "PAPER"}
    assert {offer["base_currency"] for offer in offers} == {"USD"}
    assert all(offer["collection_location_code"] for offer in offers)
    assert all(offer["collection_location_name"] for offer in offers)
    assert all(offer["destination_location_code"] for offer in offers)
    assert all(offer["destination_location_name"] for offer in offers)
    assert {
        (charge["charge_type"], charge["basis"], float(charge["amount"]), charge["currency"])
        for charge in charges
    } == {("origin", "per B/L", 30.0, "GBP")}
    assert {charge["raw_value"] for charge in charges} == {
        "USD 40 per B/L",
        "GBP 30 per B/L",
    }
    assert len(tier_tables["PEUTE"]) == 252
    assert len(tier_tables["PAPER"]) == 252
    assert {row["documentation"] for rows in tier_tables.values() for row in rows} == {
        "GBP 30 per B/L"
    }
    assert validation["summary"] == {"errors": 0, "warnings": 0, "info": 0}


def test_hapag_door_matrix_import_adds_conditional_container_charges(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    raw_dir = tmp_path / "incoming"
    raw_dir.mkdir()
    source = raw_dir / "HAPAG - FAR EAST RATES.xlsx"
    source.write_bytes(Path("rate_sheet_files/HAPAG - FAR EAST RATES.xlsx").read_bytes())
    seed_templates(tmp_path)

    result = runner.invoke(app, ["import", str(source)])
    assert result.exit_code == 0
    assert "Template used: hapag_door_matrix_v1" in result.stdout
    import_id = next(line.split(": ", 1)[1] for line in result.stdout.splitlines() if line.startswith("Import created:"))
    run_dir = tmp_path / "data" / "runs" / import_id

    offers = detail_rows(run_dir / "parsed_rate_offers.csv")
    charges = detail_rows(run_dir / "parsed_rate_charge_lines.csv")
    assert len(offers) == 546
    assert len(charges) == 1248
    assert {offer["service_mode"] for offer in offers} == {"SD / CY"}
    assert {offer["equipment_type"] for offer in offers} == {"40HC"}
    assert {offer["valid_from"] for offer in offers} == {"2026-08-01"}
    assert {offer["valid_to"] for offer in offers} == {"2026-08-31"}

    dartford = {
        offer["pod"]: offer
        for offer in offers
        if offer["place_of_receipt"] == "Dartford"
    }
    assert float(dartford["Cat Lei Terminal"]["base_amount"]) == 435.0
    assert float(dartford["Binh Duong Terminal"]["base_amount"]) == 415.0
    assert dartford["Binh Duong Terminal"]["pol"] == "London Gateway"
    assert dartford["Binh Duong Terminal"]["routing_note"] == "SGSIN-VNVUT-VNSGN"

    charges_by_offer = {}
    for charge in charges:
        charges_by_offer.setdefault(charge["rate_offer_id"], []).append(charge)
    cat_lei_charges = charges_by_offer[dartford["Cat Lei Terminal"]["id"]]
    binh_duong_charges = charges_by_offer[dartford["Binh Duong Terminal"]["id"]]
    lat_krabang_charges = charges_by_offer[dartford["Lat Krabang"]["id"]]
    assert [(charge["charge_name"], float(charge["amount"])) for charge in cat_lei_charges] == [
        ("Live Position", 15.0),
        ("Origin Docs Charges", 25.0),
    ]
    assert [(charge["charge_name"], float(charge["amount"])) for charge in binh_duong_charges] == [
        ("Live Position", 15.0),
        ("Emergency Fuel Destination", 20.0),
        ("Origin Docs Charges", 25.0),
    ]
    assert [(charge["charge_name"], float(charge["amount"])) for charge in lat_krabang_charges] == [
        ("Live Position", 15.0),
        ("Emergency Fuel Destination", 20.0),
        ("Origin Docs Charges", 25.0),
    ]

    approve_response = api_client.post(
        f"/api/imports/{import_id}/approve",
        json={
            "approved_by": "jorge",
            "carrier_name": "Hapag-Lloyd",
            "carrier_key": "hapag-door",
            "carrier_label": "Hapag-Lloyd · Door-to-quay",
        },
    )
    assert approve_response.status_code == 200
    search = api_client.get(
        "/api/search",
        params={"collection": "Dartford", "pod": "Binh Duong", "limit": 20},
    ).json()
    assert len(search) == 1
    assert search[0]["all_in_amount"] == 450.0
    assert search[0]["carrier_label"] == "Hapag-Lloyd · Door-to-quay"
    origin_docs = next(
        line
        for group in search[0]["charge_analysis"]["groups"]
        for line in group["lines"]
        if line["name"] == "Origin Docs Charges"
    )
    assert origin_docs["basis"] == "per_bill_of_lading"
    assert origin_docs["currency"] == "GBP"
    assert origin_docs["unit_amount"] == 25.0
    assert origin_docs["counts_toward_total"] is False

    desk = api_client.get("/api/rate-desk", params={"limit": 1000}).json()
    assert desk["haulage_tariffs"] == {}
    assert desk["filters"]["door_pickups"] == []
    assert "Dartford, GB" in desk["filters"]["collection_places"]
    canonical_search = api_client.get(
        "/api/rate-desk/search",
        params={"collection": "Dartford, GB", "pod": "Ho Chi Minh, VN"},
    ).json()
    canonical_rate = canonical_search["rates"][0]
    assert canonical_rate["collection_location_code"] == "dartford-gb"
    assert canonical_rate["collection_location_name"] == "Dartford, GB"
    assert canonical_rate["destination_location_code"] == "ho-chi-minh-vn"
    assert canonical_rate["destination_location_name"] == "Ho Chi Minh, VN"
    assert canonical_rate["place_of_receipt"] == "Dartford"
    assert canonical_rate["final_destination"] == "Cat Lei Terminal"


def test_hapag_india_rows_preserve_all_charges_but_total_only_selected_codes(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    raw_dir = tmp_path / "incoming"
    raw_dir.mkdir()
    source = raw_dir / "HAPAG India.xlsx"
    source.write_bytes(Path("rate_sheet_files/september rates/HAPAG India.xlsx").read_bytes())
    seed_templates(tmp_path)

    result = runner.invoke(app, ["import", str(source)])
    assert result.exit_code == 0
    assert "Parser family: hapag_india_rows" in result.stdout
    assert "Template used: hapag_india_rows_v1" in result.stdout
    import_id = next(
        line.split(": ", 1)[1]
        for line in result.stdout.splitlines()
        if line.startswith("Import created:")
    )
    run_dir = tmp_path / "data" / "runs" / import_id
    offers = detail_rows(run_dir / "parsed_rate_offers.csv")
    charges = detail_rows(run_dir / "parsed_rate_charge_lines.csv")
    assert len(offers) == 35
    assert len(charges) == 315
    assert offers[0]["place_of_receipt"] == "Norwich"
    assert offers[1]["place_of_receipt"] == "Leatherhead"
    assert offers[0]["transit_time_days"] == "45"
    assert {offer["valid_from"] for offer in offers} == {"2026-09-01"}
    assert {offer["valid_to"] for offer in offers} == {"2026-09-30"}

    first_offer_charges = [charge for charge in charges if charge["rate_offer_id"] == offers[0]["id"]]
    assert [charge["source_label"] for charge in first_offer_charges] == [
        "LUMPSUM", "LPC", "EOD", "THD", "ISF", "WHD", "DDF", "EMF", "MTD"
    ]

    approve_response = api_client.post(
        f"/api/imports/{import_id}/approve",
        json={
            "approved_by": "jorge",
            "carrier_name": "Hapag-Lloyd",
            "carrier_key": "hapag-india",
            "carrier_label": "Hapag-Lloyd · India Door-to-quay",
        },
    )
    assert approve_response.status_code == 200

    search_response = api_client.get(
        "/api/search",
        params={"collection": "Norwich", "pod": "Tuticorin", "limit": 20},
    )
    assert search_response.status_code == 200
    rows = search_response.json()
    assert len(rows) == 1
    row = rows[0]
    assert row["base_amount"] == 600.0
    assert row["charge_total"] == 85.0
    assert row["all_in_amount"] == 685.0
    assert row["all_in_usd"] == 685.0
    assert row["carrier_key"] == "hapag-india"
    analysis_lines = [line for group in row["charge_analysis"]["groups"] for line in group["lines"]]
    assert {line["name"] for line in analysis_lines} == {
        "Lumpsum", "LPC", "EOD", "THD", "ISF", "WHD", "DDF", "EMF", "MTD"
    }
    assert next(line for line in analysis_lines if line["name"] == "THD")["counts_toward_total"] is False


def test_api_import_approve_and_search_flow(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("RATE_INGEST_ROOT", str(tmp_path))
    seed_templates(tmp_path)
    source_bytes = Path("rate_sheet_files/MSC - FAR EAST RATES JAN.xlsx").read_bytes()

    response = api_client.post(
        "/api/imports",
        data={"uploaded_by": "jorge"},
        files={"file": ("MSC - FAR EAST RATES JAN.xlsx", source_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["parser_family"] == "tabular_lane"
    import_id = payload["import_id"]

    detail_response = api_client.get(f"/api/imports/{import_id}")
    assert detail_response.status_code == 200
    detail = detail_response.json()
    assert detail["summary"]["rate_offers"] > 0

    approve_response = api_client.post(
        f"/api/imports/{import_id}/approve",
        json={
            "approved_by": "jorge",
            "carrier_name": "MSC",
            "carrier_key": "msc-peute",
            "carrier_label": "MSC — PEUTE",
            "contract_tag": "PEUTE",
        },
    )
    assert approve_response.status_code == 200
    approved = approve_response.json()
    assert approved["rate_import"]["status"] == "approved"

    search_response = api_client.get("/api/search", params={"pod": "HO CHI MINH", "limit": 20})
    assert search_response.status_code == 200
    search_rows = search_response.json()
    assert search_rows
    assert any("HO CHI MINH" in (row.get("pod") or row.get("final_destination") or "") for row in search_rows)

    desk_response = api_client.get("/api/rate-desk")
    assert desk_response.status_code == 200
    desk = desk_response.json()
    assert desk["rates"]
    assert desk["last_refreshed"]
    assert desk["filters"]["origins"]
    assert desk["filters"]["destinations"]
    assert desk["filters"]["equipment_types"]
    assert "Paper" in desk["filters"]["materials"]
    assert desk["rates"][0]["source_file_name"] == "MSC - FAR EAST RATES JAN.xlsx"
    assert desk["rates"][0]["carrier_key"] == "msc-peute"

    imports_response = api_client.get("/api/imports")
    assert imports_response.status_code == 200
    listed_import = next(item for item in imports_response.json() if item["import_id"] == import_id)
    assert listed_import["carrier_label"] == "MSC — PEUTE"
    assert listed_import["lane_count"] > 0

    ui_response = api_client.get("/ui/")
    assert ui_response.status_code == 200
    assert "Reudan Rate Desk" in ui_response.text
    assert "Origin port (POL)" in ui_response.text
    assert "Material" in ui_response.text
    assert "Collection place" in ui_response.text
    assert "Carrier" in ui_response.text
    assert "Service" in ui_response.text
    assert "Routing mode" not in ui_response.text
    assert "Query API" not in ui_response.text
    assert "Import Rate File" not in ui_response.text

    import_ui_response = api_client.get("/ui/import.html")
    assert import_ui_response.status_code == 200
    assert "Drop rate sheets here" in import_ui_response.text
    assert "Review parsed sheet" in import_ui_response.text
    assert "Current file" in import_ui_response.text
    assert "Provider" in import_ui_response.text
    assert "Service" in import_ui_response.text
    assert "Query API" not in import_ui_response.text

    removed_query_ui_response = api_client.get("/ui/query-api.html")
    assert removed_query_ui_response.status_code == 404

    replacement_response = api_client.post(
        "/api/imports",
        data={"uploaded_by": "priya"},
        files={"file": ("MSC - FAR EAST RATES FEB.xlsx", source_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert replacement_response.status_code == 200
    replacement_id = replacement_response.json()["import_id"]
    replacement_approval = api_client.post(
        f"/api/imports/{replacement_id}/approve",
        json={
            "approved_by": "priya",
            "carrier_name": "MSC",
            "carrier_key": "msc-peute",
            "carrier_label": "MSC — PEUTE",
            "contract_tag": "PEUTE",
        },
    )
    assert replacement_approval.status_code == 200
    statuses = {item["import_id"]: item["status"] for item in api_client.get("/api/imports").json()}
    assert statuses[import_id] == "archived"
    assert statuses[replacement_id] == "approved"
    assert {rate["source_file_name"] for rate in api_client.get("/api/rate-desk").json()["rates"]} == {
        "MSC - FAR EAST RATES FEB.xlsx"
    }

    delete_response = api_client.delete(f"/api/imports/{replacement_id}")
    assert delete_response.status_code == 200
    assert api_client.get("/api/rate-desk").json()["rates"] == []


def detail_rows(path: Path) -> list[dict[str, str]]:
    import csv

    with path.open(encoding="utf-8") as handle:
        return list(csv.DictReader(handle))
